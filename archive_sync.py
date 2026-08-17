"""
Past seasons' tours, read from the archived BALANCE workbooks. READ-ONLY.

The 2026 season is driven by the master schedule sheet: it says when each tour
runs, and the balance workbook only adds the money. Earlier seasons have no
schedule sheet left — the balance workbook is all there is. It carries enough
on its own: column A dates every line of the itinerary, so the tour's dates,
its length and the months its service falls in can all be read straight off the
tab, and the summary block at the bottom is the same one the 2026 sheets use.

Which year a tab belongs to is decided by those dates rather than by its title:
tour codes are only MMDD, so the workbooks hold several seasons side by side.
"""
import io
import re
import datetime as _dt
from collections import Counter, defaultdict

import requests
from openpyxl import load_workbook

from profit_sync import _parse_worksheet, _usd_from_row

# The archived balance workbooks, same Drive folder as the current ones.
ARCHIVE_SHEET_IDS = {
    "TK_TV":   "1wOatDOmgZ8ri532Qas_hvRe8MYXFgNNBlm-AHoZtvU4",
    "TN_TE_M": "1hCd7l1PIafoV9-rdb0vUzqLI2MnQFSLf",
    "HM_HT_H": "1FjMTzkHtZ0KOwI5Id_x50Y-7jDossfZ8",
}

# Older seasons used series the current schedule no longer runs (TN, KN, AZ,
# TE, H, M …), and often wrote the code without its dash. Longer prefixes come
# first so "HM1016" isn't read as an "H" tour.
ARCHIVE_CODE_RE = re.compile(
    r'((?:TES|BG|TN|TE|TM|TK|TV|LT|AZ|KT|KN|HM|HT|H|M)-?\d{4}[AB]?)')


def _norm_archive_code(code: str) -> str:
    return re.sub(r'^([A-Z]+)-?(\d{4})', r'\1-\2', code.strip())


def _series_of(code: str) -> str:
    m = re.match(r'^([A-Z]+)', code)
    return m.group(1) if m else 'სხვა'


def _dates_in(ws) -> list:
    """Every date in column A — the tour's itinerary, one row per service."""
    out = []
    for row in ws.iter_rows(min_col=1, max_col=1, values_only=True):
        v = row[0]
        if isinstance(v, _dt.datetime):
            out.append(v.date())
        elif isinstance(v, _dt.date):
            out.append(v)
    return out


def _run_of_dates(dates: list) -> list:
    """The tour's own days, picked out of everything column A mentions.

    A tab usually dates one row per service, day after day, but a few also
    carry a stray date either side — a deposit paid a week early, a bill
    settled after the group left. Splitting the dates into runs and keeping the
    longest leaves the tour itself and drops those.
    """
    ordered = sorted(set(dates))
    runs, current = [], [ordered[0]]
    for prev, cur in zip(ordered, ordered[1:]):
        if (cur - prev).days <= 2:
            current.append(cur)
        else:
            runs.append(current)
            current = [cur]
    runs.append(current)
    return max(runs, key=len)


def _month_split(ws, span: list) -> list:
    """Each month's share of the tour's cost, from the dates on the lines.

    A line dated in a month counts towards that month. Tabs where the priced
    lines carry no date fall back to an even spread over the tour's days, which
    is what the 2026 sheets do when their itinerary can't place an item.
    """
    per_month = defaultdict(float)
    total = 0.0
    for row in ws.iter_rows(values_only=True):
        cells = [str(c).strip() if c is not None else '' for c in row]
        d = row[0] if row else None
        if isinstance(d, _dt.datetime):
            d = d.date()
        if not isinstance(d, _dt.date):
            continue
        usd = _usd_from_row(cells[1:])
        if usd:
            per_month[d.isoformat()[:7]] += usd
            total += usd
    if total <= 0:
        for d in span:
            per_month[d.isoformat()[:7]] += 1.0
        total = float(len(span)) or 1.0
    return [{'month': m, 'share': per_month[m] / total} for m in sorted(per_month)]


def _parse_tab(ws) -> dict:
    """One tour as the profit views want it, or None when the tab isn't one."""
    dates = _dates_in(ws)
    if not dates:
        return None
    # A tab can hold a stray date from a neighbouring season; the year that most
    # of the itinerary sits in is the tour's own.
    year = Counter(d.year for d in dates).most_common(1)[0][0]
    in_year = _run_of_dates([d for d in dates if d.year == year])
    start, end = min(in_year), max(in_year)

    code, data = _parse_worksheet(ws, ARCHIVE_CODE_RE, _norm_archive_code)
    if not code:
        # One-off charters are titled by the client's name rather than by a
        # code; keep them, since their revenue is part of the season all the
        # same. The name stands in for the code, so the same parser can run.
        name = re.sub(r'\d+\s*\+\s*\d+', '', ws.title or '').strip()
        if not name:
            return None
        code, data = _parse_worksheet(
            ws, re.compile('(' + re.escape(name) + ')'), lambda c: c)
    if not data or data.get('revenue_usd') is None:
        return None

    span_days = (end - start).days + 1
    return {
        'year': year,
        'tour_code': code,
        'series': _series_of(code),
        'pax': data['pax'],
        'rooms': data['rooms'],
        'profit_usd': data['profit_usd'],
        'vat_usd': data['vat_usd'],
        'profit_after_vat': data['profit_after_vat'],
        'spent_usd': data['spent_usd'],
        'revenue_usd': data['revenue_usd'],
        'components': data['components'],
        'components_detail': data['items'],
        'bus_start': start.isoformat(),
        'bus_end': end.isoformat(),
        'days': span_days,
        'nights': span_days - 1,
        'vat_months': _month_split(ws, in_year),
    }


def parse_workbook(content: bytes) -> list:
    wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    out = []
    for ws in wb.worksheets:
        try:
            tour = _parse_tab(ws)
        except Exception as e:
            print(f"[archive_sync] tab '{ws.title}' skipped: {e}")
            continue
        if tour:
            out.append(tour)
    wb.close()
    return out


def _download(sheet_id: str) -> bytes:
    """The workbook's bytes, however Drive happens to be holding it.

    A balance file the team built in Sheets is exported; one they uploaded as
    .xlsx is downloaded as the file it already is. Neither URL errors when it
    can't serve the file — Drive answers with an HTML page asking to sign in —
    so what comes back is checked for being a workbook (a zip, "PK") rather
    than trusted on the status code.
    """
    problems = []
    for url in (f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx",
                f"https://drive.google.com/uc?export=download&id={sheet_id}"):
        try:
            resp = requests.get(url, timeout=120, allow_redirects=True)
            resp.raise_for_status()
            if resp.content[:2] == b'PK':
                return resp.content
            problems.append(f"{url.split('//')[1][:30]}… answered "
                            f"{resp.headers.get('content-type', '?')} "
                            f"({len(resp.content)} bytes), not a workbook — "
                            f"is the file shared with anyone holding the link?")
        except Exception as e:
            problems.append(f"{url.split('//')[1][:30]}… {e}")
    raise RuntimeError('; '.join(problems))


def fetch_archive() -> dict:
    """Every tour in the archived workbooks, plus how each workbook fared.

    One workbook failing is easy to miss — the tab still fills up from the
    others — so what each one contributed is reported alongside the tours.
    """
    results = []
    files = []
    seen = set()
    for name, sheet_id in ARCHIVE_SHEET_IDS.items():
        try:
            tours = parse_workbook(_download(sheet_id))
        except Exception as e:
            print(f"[archive_sync] Could not fetch/parse {name}: {e}")
            files.append({'file': name, 'ok': False, 'tours': 0, 'error': str(e)})
            continue
        for t in tours:
            key = (t['year'], t['tour_code'])
            if key in seen:
                continue
            seen.add(key)
            results.append(t)
        by_year = {str(y): n for y, n in sorted(Counter(t['year'] for t in tours).items())}
        files.append({'file': name, 'ok': True, 'tours': len(tours), 'years': by_year})
        print(f"[archive_sync] {name}: {len(tours)} tours {by_year}")
    print(f"[archive_sync] Total: {len(results)} archived tours")
    return {'tours': results, 'files': files}


def fetch_archive_tours() -> list:
    return fetch_archive()['tours']
