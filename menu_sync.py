"""
Sync restaurant dish lists and per-dish portion ratios from the office's
Menu_2026.xlsx costing sheet. READ-ONLY.

The sheet is one tab per restaurant — a route-based restaurant like დიარონი
gets one tab per leg, and ზღაპარი gets one tab per pax/menu variant. Row 1
names the restaurant in column B; every row after that is one dish, with an
optional ratio note in column C ("N კაცზე M" — N people per M portions, or
"კაცზე N" — N portions per person) wherever the office prices it by ratio
instead of the general portion table. A "მძღოლი გიდი" row marks the start
of that tab's small side-table costing, and everything from there down is
skipped — it repeats some of the same dishes at a smaller reference
quantity, not new items.

A route-based or variant-based restaurant is folded into the same
app-facing key menu_data already uses: დიარონი's two tabs key by
(previous day's city, this day's city), exactly matching
menu_data._DIARONI_ROUTES; ზღაპარი's "TM ტური" tab keys to
"ზღაპარი (ტმ მენიუ)", the name TM's own balance sheet writes.

A couple of rules live only in office instructions, never on this sheet at
all (water's 1.15-per-2-people ratio; the "(N ადამიანზე M)" vs.
"(გაყოფილი ორად)" note wording for specific dishes) — those are folded in
from MANUAL_RATIOS / NOTE_OVERRIDES below rather than parsed.

If the workbook can't be reached or parses out empty, the caller
(menu_data.sync_menu_data) leaves whatever it already has untouched —
same "never erase good data over a bad fetch" rule the other syncs follow.
"""
import io
import re
import requests
from openpyxl import load_workbook

MENU_SHEET_ID = "1uDHSD5EAzVhlJw2rF7RV5DlFzME4HHm8"

_DRIVER_ROW_RE = re.compile(r'მძღოლი.*გიდი')
_NUMERIC_ROW_RE = re.compile(r'^[\d.,%\s]+$')
_RATIO_DIVIDE_RE = re.compile(r'^(\d+)\s*კაცზე\s*(\d+)$')      # "2 კაცზე 1" -> 2 people : 1 portion
_RATIO_MULTIPLY_RE = re.compile(r'^კაცზე\s*(\d+)$')             # "კაცზე 2" -> 2 portions per person
_RATIO_LITERAL_DIVIDE_RE = re.compile(r'^(\d+)\s*ადამიანზე\s*(\d+)$')  # already-worded "2 ადამიანზე 1"

# The office has started writing the finished note text straight into the
# ratio cell for some dishes, instead of the raw "N კაცზე M" — recognized
# as itself rather than re-derived.
_LITERAL_NOTES = {"გაყოფილი ორად"}

# Business rules not written anywhere on this sheet.
MANUAL_RATIOS = {
    (None, "წყალი"): (1.15, 2, None),
}
# How to word a divide-ratio's note for specific dishes — anything not
# listed here just gets the sheet's own wording, "N ადამიანზე M".
NOTE_OVERRIDES = {
    "სუფი": "გაყოფილი ორად",
    "სოკოს სუპი": "გაყოფილი ორად",
    "ბოსტნეულის სუპი": "გაყოფილი ორად",
}


def _parse_ratio(text: str):
    """Ratio cell text -> (numerator, denominator, note), or None.

    The 1:2 divide ratio can show up three ways depending on how this
    particular tab was last edited: the raw "2 კაცზე 1", the same thing
    already worded as "2 ადამიანზე 1", or the office's own finished phrase
    for it, "გაყოფილი ორად" — all three mean the same 1-portion-per-2-people
    ratio, just spelled out to a different degree.
    """
    text = (text or "").strip()
    if not text:
        return None
    if text in _LITERAL_NOTES:
        return 1, 2, text
    m = _RATIO_DIVIDE_RE.match(text) or _RATIO_LITERAL_DIVIDE_RE.match(text)
    if m:
        people, portions = int(m.group(1)), int(m.group(2))
        return portions, people, f"{people} ადამიანზე {portions}"
    m = _RATIO_MULTIPLY_RE.match(text)
    if m:
        return int(m.group(1)), 1, None
    return None


def _restaurant_key_from_title(title: str):
    """(app-facing key, route) for one worksheet, from its tab title.

    The tab title is the reliable signal — row 1's own restaurant-name cell
    is inconsistent between tabs (one დიარონი route tab spells it
    "დიარონი: <route>" in columns B/C, the other just puts the route phrase
    straight in column B with no restaurant name at all), but every tab's
    title itself is consistently "<restaurant> [variant]"."""
    title = (title or '').strip()
    if title.startswith('დიარონი'):
        rest = title[len('დიარონი'):].strip()
        if 'ბათუმი' in rest and 'მესტ' in rest:
            return 'დიარონი', ('Batumi', 'Mestia')
        if 'მესტ' in rest and 'გორ' in rest:
            return 'დიარონი', ('Mestia', 'Gori')
        return 'დიარონი', None
    if title.startswith('ზღაპარი'):
        if 'tm' in title.lower():
            return 'ზღაპარი (ტმ მენიუ)', None
        return 'ზღაპარი', None
    return title, None


def _parse_tab(ws):
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return None
    key, route = _restaurant_key_from_title(ws.title)
    if not key:
        return None

    dishes = []
    seen = set()
    past_driver_row = False
    for row in rows[1:]:
        cell_b = str(row[1] or '').strip() if len(row) > 1 else ''
        if not cell_b:
            continue
        if _DRIVER_ROW_RE.search(cell_b):
            past_driver_row = True
            continue
        if past_driver_row or _NUMERIC_ROW_RE.match(cell_b) or cell_b in seen:
            continue
        seen.add(cell_b)
        # The ratio cell's exact column isn't consistent between tabs (one
        # tab leaves an extra blank column before it) — scan the rest of the
        # row for whichever cell actually matches the ratio pattern, rather
        # than trusting a fixed position.
        ratio = None
        for cell in row[2:]:
            ratio = _parse_ratio(str(cell or '').strip())
            if ratio:
                break
        if ratio:
            num, den, note = ratio
            note = NOTE_OVERRIDES.get(cell_b, note)
            ratio = (num, den, note)
        dishes.append({"name": cell_b, "ratio": ratio})
    return {"key": key, "route": route, "dishes": dishes}


def _parse_workbook(content: bytes) -> dict:
    restaurants: dict = {}
    routes: dict = {}
    ratios: dict = dict(MANUAL_RATIOS)
    wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    for ws in wb.worksheets:
        parsed = _parse_tab(ws)
        if not parsed or not parsed["dishes"]:
            continue
        key, route = parsed["key"], parsed["route"]
        names = [d["name"] for d in parsed["dishes"]]
        if route:
            routes.setdefault(key, {})[route] = names
            # Also fold into the bare key's own list, as the fallback for
            # when a day's route can't be worked out.
            existing = restaurants.setdefault(key, [])
            for n in names:
                if n not in existing:
                    existing.append(n)
        else:
            restaurants[key] = names
        for d in parsed["dishes"]:
            if d["ratio"]:
                ratios[(key, d["name"])] = d["ratio"]
    wb.close()
    return {"restaurants": restaurants, "routes": routes, "ratios": ratios}


def fetch_menu() -> dict:
    """Return {"restaurants": {key: [dish,...]}, "routes": {key: {(prev,cur): [dish,...]}},
    "ratios": {(restaurant_or_None, dish): (num,den,note)}}, or {} on failure."""
    try:
        url = f"https://docs.google.com/spreadsheets/d/{MENU_SHEET_ID}/export?format=xlsx"
        resp = requests.get(url, timeout=60)
        resp.raise_for_status()
        result = _parse_workbook(resp.content)
        print(f"[menu_sync] parsed {len(result['restaurants'])} restaurants, "
              f"{sum(len(v) for v in result['routes'].values())} route variants, "
              f"{len(result['ratios'])} ratios")
        return result
    except Exception as e:
        print(f"[menu_sync] Could not fetch/parse Menu_2026.xlsx: {e}")
        return {}
