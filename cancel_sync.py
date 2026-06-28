"""
Sync cancelled tours from the master schedule Google Sheet's cancelled tab
(gid=1709702486).  Any tour code found there must be removed from the local DB.

READ-ONLY on the sheet — never write.
"""
import io
import re
import requests
from openpyxl import load_workbook

MASTER_SHEET_ID = "13FoSFZqpi4QAm2CDc1qT3uB7AKHOFEJv"

TOUR_CODE_RE = re.compile(r'\b((?:ZT|LN|KT|DT1|DT2|LT|ST|MT|HM)-?\d{4})\b')


def _norm_code(code: str) -> str:
    return re.sub(r'(ZT|LN|KT|DT1|DT2|LT|ST|MT|HM)(\d{4})', r'\1-\2', code)


def fetch_cancelled_tour_codes() -> set:
    """Return a set of normalised tour codes from the cancelled tab."""
    url = f"https://docs.google.com/spreadsheets/d/{MASTER_SHEET_ID}/export?format=xlsx"
    try:
        resp = requests.get(url, timeout=60)
        resp.raise_for_status()
        wb = load_workbook(io.BytesIO(resp.content), data_only=True, read_only=True)

        target_ws = None

        # Primary: find sheet whose title contains "cancel"
        for ws in wb.worksheets:
            if 'cancel' in (ws.title or '').lower():
                target_ws = ws
                break

        # Fallback: first sheet that has "for cancell" in its first 10 rows
        if target_ws is None:
            for ws in wb.worksheets:
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    if i >= 10:
                        break
                    row_text = ' '.join(str(c).lower() for c in row if c)
                    if 'for cancell' in row_text:
                        target_ws = ws
                        break
                if target_ws:
                    break

        if target_ws is None:
            print("[cancel_sync] Cancelled tab not found in workbook")
            wb.close()
            return set()

        cancelled: set = set()
        for row in target_ws.iter_rows(values_only=True):
            for cell in row:
                if not cell:
                    continue
                for m in TOUR_CODE_RE.finditer(str(cell)):
                    cancelled.add(_norm_code(m.group(1)))

        wb.close()
        print(f"[cancel_sync] Found {len(cancelled)} cancelled codes: {sorted(cancelled)}")
        return cancelled

    except Exception as e:
        print(f"[cancel_sync] Could not fetch/parse master sheet: {e}")
        return set()
