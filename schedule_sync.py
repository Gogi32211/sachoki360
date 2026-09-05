"""
Sync the current + planned tour schedule from the master Google Sheet
"TOURS_All Dates_2026". READ-ONLY.

The sheet lists every tour as a column block. The region BEFORE the
"done 2026" marker holds ongoing + planned tours ("IN PROCESS"); after it
come completed tours, and further down the cancelled ("for cancell")
sections. We only treat the pre-"done 2026" region as the active list.

A tour code encodes its first (Baku/Almaty) date as MMDD; the app's
bus_start is that date plus a per-series offset (the Georgia / Khareba day).
"""
import io
import re
from datetime import date, timedelta
import requests
from openpyxl import load_workbook

from seed_data import SERIES_START_OFFSET

SHEET_ID = "13FoSFZqpi4QAm2CDc1qT3uB7AKHOFEJv"

TOUR_CODE_RE = re.compile(r'\b((?:ZT|LN|KT|DT1|DT2|LT|HM1|HM2|HM|HT|TH|TK|TM|TV|MT|ST)-?\d{4})\b')
DONE_MARKER = "done 2026"
_ROOM_KEYWORD_RE = re.compile(
    r'\d+\s*(?:twin|single|double|king|suite)\b|\b(?:twin|single|double|king|suite)\b',
    re.IGNORECASE)
_ROOM_ENTRY_RE = re.compile(r'(\d+)\s*(twin|single|double|king|suite)\s*(?:\(([^)]*)\))?', re.IGNORECASE)
_DATE_LIKE_RE = re.compile(
    r'^\d{1,2}/\d{1,2}/\d{2,4}$|^\d{4}-\d{2}-\d{2}(\s+\d{2}:\d{2}:\d{2})?$')
_DRIVER_RE = re.compile(r'^([\d][\d\s\-]{5,})\s+(\S.*)$')


def _norm_code(code: str) -> str:
    return re.sub(r'(ZT|LN|KT|DT1|DT2|LT|HM1|HM2|HM|HT|TH|TK|TM|TV|MT|ST)(\d{4})', r'\1-\2', code)


def _bus_start_from_code(code: str, series: str):
    """code SERIES-MMDD → date(2026, MM, DD) + series offset → ISO bus_start."""
    m = re.search(r'-(\d{2})(\d{2})$', code)
    if not m:
        return None
    off = SERIES_START_OFFSET.get(series)
    if off is None:
        return None
    try:
        d = date(2026, int(m.group(1)), int(m.group(2))) + timedelta(days=off)
        return d.isoformat()
    except ValueError:
        return None


def _abbrev_rooms(text: str) -> str:
    """Convert room description to abbreviated form.

    Examples:
      "6 Twin + 1 Twin(one guest) + 1 Single(leader)"  →  "6T+1TG+1SL"
      "9TWIN,2SINGLE/19+1"                              →  "9T+2S"
      "21+1/10TWIN,1DOUBLE"                             →  "10T+1D"
      "11T"                                             →  "11T"
    """
    if not text:
        return ''
    parts = []
    for m in _ROOM_ENTRY_RE.finditer(text):
        count = m.group(1)
        rtype = m.group(2).lower()
        qual  = (m.group(3) or '').lower()
        if rtype == 'twin':
            abbr = f"{count}TG" if 'guest' in qual else f"{count}T"
        elif rtype == 'single':
            if 'leader' in qual:
                abbr = f"{count}SL"
            elif 'guest' in qual:
                abbr = f"{count}SG"
            else:
                abbr = f"{count}S"
        elif rtype == 'double':
            abbr = f"{count}D"
        elif rtype == 'king':
            abbr = f"{count}K"
        else:
            abbr = f"{count}?"
        parts.append(abbr)
    return '+'.join(parts)


def fetch_active_tours() -> list:
    """Return [{code, series, bus_start, rooms}] for ongoing/planned tours."""
    url = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=xlsx"
    resp = requests.get(url, timeout=60)
    resp.raise_for_status()
    wb = load_workbook(io.BytesIO(resp.content), data_only=True, read_only=True)

    # Only read the main schedule tab — reading all tabs would pull tour codes
    # from the 'cancelled' tab and incorrectly treat them as active.
    main_ws = None
    for ws in wb.worksheets:
        title = (ws.title or '').lower()
        if '2026' in title or 'tour' in title or 'map' in title:
            main_ws = ws
            break
    if main_ws is None:
        main_ws = wb.worksheets[0]

    # Read as 2-D grid to preserve column positions for rooms look-up.
    grid = []
    for row in main_ws.iter_rows(values_only=True):
        grid.append([str(c).strip() if c is not None else '' for c in row])
    wb.close()

    # Find the "done 2026" boundary row — only scan above it.
    done_row = len(grid)
    for i, row in enumerate(grid):
        if any(DONE_MARKER in c.lower() for c in row):
            done_row = i
            break

    seen: set = set()
    active: list = []

    for row_i in range(done_row):
        for col_i, cell in enumerate(grid[row_i]):
            m = TOUR_CODE_RE.search(cell)
            if not m:
                continue
            code = _norm_code(m.group(1))
            if code in seen:
                continue
            seen.add(code)
            series = code.split('-')[0]
            bs = _bus_start_from_code(code, series)
            if not bs:
                continue

            # Look up to 4 rows above in the same column for a rooms/pax row.
            rooms = ''
            for look_back in range(1, 5):
                ri = row_i - look_back
                if ri < 0:
                    break
                candidate = grid[ri][col_i] if col_i < len(grid[ri]) else ''
                if _ROOM_KEYWORD_RE.search(candidate):
                    rooms = _abbrev_rooms(candidate)
                    break

            active.append({"code": code, "series": series, "bus_start": bs,
                           "rooms": rooms, "guide": _guide_above(grid, row_i, col_i)})

    print(f"[schedule_sync] active tours parsed: {len(active)}")
    return active


# The guide's name sits a row or two above the tour code in the same column,
# alongside the rooms line — e.g. "IM danieli" over "10 twin/19+1" over "ZT-0803".
# There is no marker to key on, so identify it by ruling out everything else that
# appears there: room specs, dates, counts, tour codes and section banners.
_GUIDE_SKIP_RE = re.compile(
    r'(in process|done|cancel|for cancell|ok\b|paid|revised)', re.IGNORECASE)
_LETTER_RE = re.compile(r'[A-Za-z\u10A0-\u10FF]')
# Deliberately looser than _ROOM_KEYWORD_RE: for ruling a cell OUT, any mention
# of a room type is enough, however it is written.
_ROOM_WORD_RE = re.compile(r'(twin|single|double|king|suite)', re.IGNORECASE)
_NUMERIC_ONLY_RE = re.compile(r'^[\d/.\-\s:+,()]+$')


def _looks_like_guide(text: str) -> bool:
    s = (text or '').strip()
    if not s or len(s) > 60:
        return False
    if TOUR_CODE_RE.search(s) or _ROOM_WORD_RE.search(s):
        return False
    if _GUIDE_SKIP_RE.search(s) or _NUMERIC_ONLY_RE.match(s):
        return False
    return bool(_LETTER_RE.search(s))


def _guide_above(grid, row_i, col_i) -> str:
    """Nearest plausible guide name in the rows just above a tour code."""
    for look_back in range(1, 6):
        ri = row_i - look_back
        if ri < 0:
            break
        candidate = grid[ri][col_i] if col_i < len(grid[ri]) else ''
        if _looks_like_guide(candidate):
            return candidate.strip()
    return ''


def _driver_below(grid, row_i, col_i) -> str:
    """Bus driver's phone + name, in the row right after the tour's last dated
    row — e.g. "598 59 19 10  გოგიტა" straight under the last date cell."""
    ri = row_i + 1
    n = len(grid)
    while ri < n:
        cell = (grid[ri][col_i] if col_i < len(grid[ri]) else '').strip()
        if not _DATE_LIKE_RE.match(cell):
            break
        ri += 1
    if ri >= n:
        return ''
    cell = (grid[ri][col_i] if col_i < len(grid[ri]) else '').strip()
    m = _DRIVER_RE.match(cell)
    if m and _LETTER_RE.search(m.group(2)):
        return cell
    return ''


def fetch_all_tour_rooms() -> dict:
    """Return {tour_code: {"rooms": str, "guide": str, "driver": str}} for ALL
    tour codes in the main tab, including completed ones past the
    'done 2026' marker."""
    url = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=xlsx"
    resp = requests.get(url, timeout=60)
    resp.raise_for_status()
    wb = load_workbook(io.BytesIO(resp.content), data_only=True, read_only=True)

    main_ws = None
    for ws in wb.worksheets:
        title = (ws.title or '').lower()
        if '2026' in title or 'tour' in title or 'map' in title:
            main_ws = ws
            break
    if main_ws is None:
        main_ws = wb.worksheets[0]

    grid = []
    for row in main_ws.iter_rows(values_only=True):
        grid.append([str(c).strip() if c is not None else '' for c in row])
    wb.close()

    seen: set = set()
    rooms_map: dict = {}

    for row_i in range(len(grid)):
        for col_i, cell in enumerate(grid[row_i]):
            m = TOUR_CODE_RE.search(cell)
            if not m:
                continue
            code = _norm_code(m.group(1))
            if code in seen:
                continue
            seen.add(code)

            rooms = ''
            for look_back in range(1, 5):
                ri = row_i - look_back
                if ri < 0:
                    break
                candidate = grid[ri][col_i] if col_i < len(grid[ri]) else ''
                if _ROOM_KEYWORD_RE.search(candidate):
                    rooms = _abbrev_rooms(candidate)
                    break

            guide = _guide_above(grid, row_i, col_i)
            driver = _driver_below(grid, row_i, col_i)
            if rooms or guide or driver:
                rooms_map[code] = {"rooms": rooms, "guide": guide, "driver": driver}

    n_rooms = sum(1 for v in rooms_map.values() if v["rooms"])
    n_guide = sum(1 for v in rooms_map.values() if v["guide"])
    n_driver = sum(1 for v in rooms_map.values() if v["driver"])
    print(f"[schedule_sync] full sheet: rooms for {n_rooms}, guides for {n_guide}, "
          f"drivers for {n_driver} tours")
    return rooms_map
