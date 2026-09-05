"""
Sync guide / hotel / restaurant phone numbers from the master schedule
workbook's "informations" tab. READ-ONLY.

Matching each phone number back to the app's own records is best-effort,
since the three lists don't share a key with what's already stored:

  - restaurants are already Georgian script, same as menu_data and
    tour_meals — near-exact, with a small alias table for known spelling
    variants between the two sheets.
  - hotels are Georgian brand names, bridged to the canonical English names
    already used in daily_log.hotel via a small hand-built table.
  - guides are Georgian names here, but tours.guide is transliterated to
    Latin by schedule_sync's sheet reader — matched by transliterating the
    Georgian name and comparing normalized words, so this one can
    occasionally miss on an unusual spelling.
"""
import io
import re
import requests
from openpyxl import load_workbook

from schedule_sync import SHEET_ID

INFO_TAB = "informations"
_PLACEHOLDER_NAMES = {"გიდი"}  # unnamed rows in the guide column

_GEO_LAT = {
    'ა': 'a', 'ბ': 'b', 'გ': 'g', 'დ': 'd', 'ე': 'e', 'ვ': 'v', 'ზ': 'z',
    'თ': 't', 'ი': 'i', 'კ': 'k', 'ლ': 'l', 'მ': 'm', 'ნ': 'n', 'ო': 'o',
    'პ': 'p', 'ჟ': 'zh', 'რ': 'r', 'ს': 's', 'ტ': 't', 'უ': 'u', 'ფ': 'p',
    'ქ': 'k', 'ღ': 'gh', 'ყ': 'q', 'შ': 'sh', 'ჩ': 'ch', 'ც': 'ts',
    'ძ': 'dz', 'წ': 'ts', 'ჭ': 'ch', 'ხ': 'kh', 'ჯ': 'j', 'ჰ': 'h',
}

# Georgian brand name (informations tab) -> canonical English name used in
# daily_log.hotel / sheets_sync.HOTEL_MAP.
HOTEL_ALIASES = {
    'ჰუალინგი': 'Hualing Tbilisi',
    'მარკო პოლო': 'Marco Polo Gudauri',
    'პულმან თბილისი': 'Pullman Tbilisi',
    'გრინვუდ ბათუმი': 'Greenwood Batumi',
    'ლილატ მესტია': 'Lilate Mestia',
    'უშბა ჰოტელ': 'Ushba Hotel',
    'გუდაური ინნ': 'Gudauri Inn',
    'გორი ინნ': 'Gori Inn',
}

# Known spelling variants between the "informations" tab and
# menu_data.RESTAURANT_MENUS / tour_meals.restaurant.
RESTAURANT_ALIASES = {
    'ლუზიასთან': 'ლუიზასთან',
    'კტვ პატარძეული': 'კტვ',
}


def _translit(s: str) -> str:
    return ''.join(_GEO_LAT.get(ch, ch) for ch in (s or '').lower())


def _words(s: str) -> list:
    s = re.sub(r'[^a-z\s]', ' ', _translit(s))
    return [w for w in s.split() if len(w) > 2]


def _norm_phone(v) -> str:
    s = str(v or '').strip()
    if s.endswith('.0'):
        s = s[:-2]
    s = re.sub(r'[ \s\-]+', ' ', s).strip()
    return s


def fetch_contacts() -> dict:
    """Return {"guides": [{"name","phone"}], "hotels": {name: {phone,phone2}},
    "restaurants": {name: {phone,phone2}}}."""
    url = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=xlsx"
    guides, hotels, restaurants = [], {}, {}
    try:
        resp = requests.get(url, timeout=60)
        resp.raise_for_status()
        wb = load_workbook(io.BytesIO(resp.content), data_only=True, read_only=True)
        ws = None
        for w in wb.worksheets:
            if (w.title or '').strip().lower() == INFO_TAB:
                ws = w
                break
        if ws is None:
            print(f"[contacts_sync] no '{INFO_TAB}' tab found")
            wb.close()
            return {"guides": [], "hotels": {}, "restaurants": {}}

        for row in ws.iter_rows(min_row=3, values_only=True):
            cells = list(row) + [None] * 11

            g_name = str(cells[1] or '').strip()
            if g_name and g_name not in _PLACEHOLDER_NAMES:
                phone = _norm_phone(cells[2])
                if phone:
                    guides.append({"name": g_name, "phone": phone})

            h_name = str(cells[4] or '').strip()
            if h_name:
                canon = HOTEL_ALIASES.get(h_name, h_name)
                hotels[canon] = {"phone": _norm_phone(cells[5]),
                                  "phone2": _norm_phone(cells[6])}

            r_name = str(cells[8] or '').strip()
            if r_name:
                canon = RESTAURANT_ALIASES.get(r_name, r_name)
                restaurants[canon] = {"phone": _norm_phone(cells[9]),
                                       "phone2": _norm_phone(cells[10])}
        wb.close()
        print(f"[contacts_sync] guides={len(guides)} hotels={len(hotels)} "
              f"restaurants={len(restaurants)}")
    except Exception as e:
        print(f"[contacts_sync] Could not fetch/parse informations tab: {e}")
    return {"guides": guides, "hotels": hotels, "restaurants": restaurants}


def match_guide_phone(guide_field: str, guides: list) -> str:
    """Best-effort: tours.guide is Latin-transliterated, the sheet's names
    are Georgian — compare transliterated, normalized word sets and take
    the best overlap. Returns '' when nothing plausible matches."""
    if not guide_field or not guides:
        return ''
    field_words = set(_words(guide_field))
    if not field_words:
        return ''
    best_phone, best_score = '', 0
    for g in guides:
        overlap = field_words & set(_words(g['name']))
        if len(overlap) > best_score:
            best_score, best_phone = len(overlap), g['phone']
    return best_phone
