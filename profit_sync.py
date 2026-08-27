"""
Sync estimated tour profit (USD) from the Google Sheets BALANCE files.

Each financial workbook has one tab per tour. Every tab ends with a summary
block. The relevant cells (all amounts in USD):

    დღგ სავარაუდო   <vat_usd>            (estimated VAT to be refunded)
    დაიხარჯა        <spent_usd>  ...
    ტურის ღირებ.    <revenue_usd> ...
    მოგება          <profit_usd>  <ignore>  <profit_after_vat>

profit_after_vat == profit_usd + vat_usd.

READ-ONLY on the sheets.
"""
import io
import re
from datetime import date

import requests
from openpyxl import load_workbook

# Same three financial workbooks used for meals/payments.
SHEET_IDS = {
    "KT_DT": "16NWhGGHR7mXAwRyVH_vmYSrZHx1zxrAR",
    "LN":    "1p5rgt6w_1hGpDr2W3Mug1p7rYWi7L7ZR",
    "ZT":    "1aWUi7GuMFZLuSq1dp2MgP_KV4rmwXGAE",
    "MT_ST": "1bzsKKc6lHIDuoeuK1WCbK_lG1mbPlkYN",
    "TM":    "1I_mMGVWcel93pNH72fYVM6sYlxOrPj0HHQiSxq2pS_o",
}

# Workbooks whose tours belong in the money but not in the schedule: another
# desk runs them, so they must not turn up in the day view or the timeline.
FINANCE_ONLY = ("MT_ST",)

TOUR_CODE_RE = re.compile(r'((?:ZT|LN|KT|DT1|DT2|LT|HM1|HM2|HM|HT|TH|TK|TM|TV|MT|ST)-?\d{4})')
# Group size, e.g. "19+1" = 19 tourists + 1 leader.
PAX_RE = re.compile(r'(?<!\d)(\d{1,2})\s*\+\s*(\d{1,2})')
_ROOM_KEYWORD_RE = re.compile(
    r'\d+\s*(?:twin|single|double|king|suite)\b|\b(?:twin|single|double|king|suite)\b',
    re.IGNORECASE)
_ROOM_ENTRY_RE = re.compile(r'(\d+)\s*(twin|single|double|king|suite)\s*(?:\(([^)]*)\))?', re.IGNORECASE)


def _extract_pax(text: str, code_re=None):
    # Tab titles concatenate code+pax (e.g. "LT-070811+1" = LT-0708, 11+1):
    # strip the tour code first so its digits can't bleed into the pax match.
    cleaned = (code_re or TOUR_CODE_RE).sub(' ', text or '')
    m = PAX_RE.search(cleaned)
    return f"{m.group(1)}+{m.group(2)}" if m else None


def _abbrev_rooms(text: str) -> str:
    if not text:
        return ''
    parts = []
    for m in _ROOM_ENTRY_RE.finditer(text):
        count = m.group(1)
        rtype = m.group(2).lower()
        qual  = (m.group(3) or '').lower()
        if rtype == 'twin':
            abbr = f"{count}TG" if 'guest' in qual else f"{count}T"
        elif rtype == 'single':
            if 'leader' in qual:
                abbr = f"{count}SL"
            elif 'guest' in qual:
                abbr = f"{count}SG"
            else:
                abbr = f"{count}S"
        elif rtype == 'double':
            abbr = f"{count}D"
        elif rtype == 'king':
            abbr = f"{count}K"
        else:
            abbr = f"{count}?"
        parts.append(abbr)
    return '+'.join(parts)


# Summary-block labels — not line items.
_SUMMARY_KEYS = ('დღგ სავარაუდო', 'დაიხარჯა', 'ტურის ღირებ', 'მოგება',
                 'ჩარიცხული', 'ჩასარიცხია', 'zedmeti')


def _usd_from_row(cells) -> float:
    """Best-effort USD amount for a line item. `cells` is the row from column A.

    A GEL/USD pair shows up as adjacent numbers with ratio ~2.4–2.95; otherwise
    the amount is entered directly in USD, always in column E — the same slot
    a paired row keeps its own USD figure in, one place left of the GEL total.
    A handful of rows carry a stray unrelated number further along the row (a
    leftover note, a different currency), so scanning the row for the largest
    number picks that up instead; column E doesn't have that problem."""
    nums = [n for n in (_num(c) for c in cells) if n is not None]
    for i in range(len(nums) - 1):
        a, b = nums[i], nums[i + 1]
        if b and 2.4 <= a / b <= 2.95:
            return b
    e = _num(cells[4]) if len(cells) > 4 else None
    if e is not None and e >= 5:
        return e
    cand = [n for n in nums if n >= 5]
    return max(cand) if cand else 0.0


# Known spelling corrections in balance sheets: wrong → correct.
_SPELLING_FIXES = {
    'ორქოს': 'ოქროს',
}


def _fix_name(name: str) -> str:
    for wrong, right in _SPELLING_FIXES.items():
        name = name.replace(wrong, right)
    return name


_HOTEL_KEYS = ('ჰუალინგ', 'radisson', 'რადისონ', 'მარკო', 'გორი ინ',
               'გრინვუდ', 'ახალციხ', 'უშბა', 'გისტოლა', 'პულმან', 'pullman',
               'ლილატ', 'გუდაურ', 'best western', 'ბესთ', 'ქრაუნ',
               'crown', 'ბორჯომ', 'ინნ', 'pine', 'პაინ', 'პინო', 'pino', 'ჯინო',
               'yerevan', 'ერევან', 'aghababayan', 'agababayan', 'სევან')


def _categorize(name: str):
    """Map a line-item description to a cost component (or None to skip)."""
    n = name.lower().strip()
    if not n:
        return None
    is_meal = any(k in n for k in ('ლანჩი', 'ვახშამ', 'ვაშამ'))
    # Wine/food tastings are an activity, not a meal → counted as an attraction.
    is_tasting = 'დეგუსტაცი' in n
    if any(k in n for k in ('სომხეთი', 'აღაბაბაია')):
        # Meal rows in Armenia are informational diary entries; the cost is
        # already captured in the guide and hotel line items.
        if is_meal or is_tasting:
            return None
        return 'armenia'
    if 'აზერბაიჯ' in n:
        # Same as Armenia: a diary meal row here is already paid for in the
        # bundled guide/hotel/bus line, not a cost of its own.
        if is_meal or is_tasting:
            return None
        return 'azerbaijan'
    if 'ავტობუს' in n or 'სპრინტერ' in n:
        return 'bus'
    is_guide_ref  = 'გიდ' in n
    is_driver_ref = 'მძღოლ' in n
    is_dito_ref   = 'დიტო' in n
    # Staff (personnel) expenses: guide/driver accommodation, transport, food
    if (is_guide_ref or is_driver_ref) and \
       any(k in n for k in ('კვება', 'დარჩენა', 'ტრანსპორტ', 'ტაქს', 'ხარჯ')):
        return 'staff'
    # "დიტო" entries in hotel rows = staff accommodation/meals at a hotel
    if is_dito_ref and any(k in n for k in ('დარჩენა', 'კვება', 'ღამ')):
        return 'staff'
    # Driver name / tips entry → merged into bus
    if is_driver_ref:
        return 'bus'
    # A tasting is an activity even when the row is worded as a meal
    # (e.g. "ვახშამი: კტვ დეგუსტაცია") — bill it to attractions, not restaurants.
    if is_tasting:
        return 'attraction'
    # Dinner/lunch at a hotel is already included in the hotel cost — skip.
    if is_meal and any(k in n for k in _HOTEL_KEYS):
        return None
    if is_meal:
        return 'restaurant'
    if is_guide_ref:
        return 'guide'
    if any(k in n for k in ('ვარძი', 'სტალინ', 'უფლისციხ', 'უფლიციხ', 'საბაგირ',
                            'გემზე', 'მოზეომ', 'დელიკ', 'ბილეთ', 'მუზეუმ', 'ცაგვ',
                            'მარტვილ', 'ბორჯომის პარკ')):
        return 'attraction'
    if any(k in n for k in _HOTEL_KEYS):
        return 'hotel'
    return 'other'


_DATE_RE = re.compile(r'(\d{4}-\d{2}-\d{2})')


def _run_of_dates(dates: list) -> list:
    """The tour's own days, picked out of everything column A mentions.

    A tab usually dates one row per service, day after day, but some also carry
    a stray date either side — a deposit paid a week early, a bill settled
    after the group left. Splitting the dates into runs and keeping the longest
    leaves the tour itself and drops those.
    """
    ordered = sorted(set(dates))
    if not ordered:
        return []
    runs, current = [], [ordered[0]]
    for prev, cur in zip(ordered, ordered[1:]):
        if (date.fromisoformat(cur) - date.fromisoformat(prev)).days <= 2:
            current.append(cur)
        else:
            runs.append(current)
            current = [cur]
    runs.append(current)
    return max(runs, key=len)


def _norm_code(code: str) -> str:
    return re.sub(r'(ZT|LN|KT|DT1|DT2|LT|HM1|HM2|HM|HT|TH|TK|TM|TV|MT|ST)(\d{4})', r'\1-\2', code)


def _num(v):
    if v is None:
        return None
    s = str(v).replace(',', '').strip()
    if s == '':
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _parse_worksheet(ws, code_re=None, norm=None) -> tuple:
    """Return (tour_code, profit_dict) or (None, None) if no code found.

    Earlier seasons number their tours differently, so the caller can hand in
    its own code pattern; everything below the code is the same workbook
    template either way.
    """
    code_re = code_re or TOUR_CODE_RE
    norm = norm or _norm_code
    code = None
    m = code_re.search(ws.title or '')
    if m:
        code = norm(m.group(1))
    pax = _extract_pax(ws.title, code_re)

    out = {
        'pax': pax,               # group size, e.g. "19+1"
        'rooms': '',              # abbreviated room config, e.g. "8T+1S+1SL"
        'profit_usd': None,       # მოგება (1st value = USD)
        'vat_usd': None,          # დღგ სავარაუდო (USD VAT to refund)
        'profit_after_vat': None, # მოგება 3rd value = profit + vat
        'spent_usd': None,        # დაიხარჯა (1st value = USD)
        'revenue_usd': None,      # ტურის ღირებ. (1st value = USD)
        'components': {},         # {category: usd} cost breakdown (no Azerbaijan)
    }
    comp = {}
    items = []  # [{name, cat, usd}] — individual line items
    dates = []  # column A dates: the tab's own itinerary

    for row in ws.iter_rows(values_only=True):
        cells = [str(c).strip() if c is not None else '' for c in row]
        d = _DATE_RE.match(cells[0]) if cells else None
        if d:
            dates.append(d.group(1))
        if not code:
            for cell in cells[:4]:
                mm = code_re.search(cell)
                if mm:
                    code = norm(mm.group(1))
                    # Also scan the same row for room configuration
                    for rc in cells:
                        if _ROOM_KEYWORD_RE.search(rc):
                            abbr = _abbrev_rooms(rc)
                            if abbr:
                                out['rooms'] = abbr
                                break
                    break
        if not out['pax']:
            for cell in cells[:4]:
                p = _extract_pax(cell, code_re)
                if p:
                    out['pax'] = p
                    break
        # Pick up rooms from any row that has room keywords (e.g. header row)
        if not out['rooms']:
            for cell in cells:
                if _ROOM_KEYWORD_RE.search(cell) and not any(k in cell.lower() for k in _SUMMARY_KEYS):
                    abbr = _abbrev_rooms(cell)
                    if abbr:
                        out['rooms'] = abbr
                        break

        # ── Cost-component breakdown from line items ──
        name = _fix_name(cells[1].strip() if len(cells) > 1 else '')
        if name and not any(name.startswith(k) for k in _SUMMARY_KEYS) \
                and not code_re.search(name):
            cat = _categorize(name)
            if cat:
                usd = _usd_from_row(cells)
                if usd:
                    comp[cat] = round(comp.get(cat, 0.0) + usd, 2)
                    items.append({'name': name, 'cat': cat, 'usd': round(usd, 2)})

        for i, c in enumerate(cells):
            is_vat = c.startswith('დღგ სავარაუდო')
            is_spent = c == 'დაიხარჯა'
            is_rev = c.startswith('ტურის ღირებ')
            is_profit = c == 'მოგება'
            if not (is_vat or is_spent or is_rev or is_profit):
                continue
            nxt = [n for n in (_num(x) for x in cells[i + 1:] if str(x).strip() != '')
                   if n is not None]
            if is_vat and len(nxt) >= 1:
                out['vat_usd'] = nxt[0]
            elif is_spent and len(nxt) >= 1:
                out['spent_usd'] = nxt[0]
            elif is_rev and len(nxt) >= 1:
                out['revenue_usd'] = nxt[0]
            elif is_profit and len(nxt) >= 1:
                out['profit_usd'] = nxt[0]
                if len(nxt) >= 3:
                    out['profit_after_vat'] = nxt[2]

    if not code:
        return None, None
    out['components'] = comp
    out['items'] = items
    # A tour the schedule sheet doesn't carry still knows when it ran: the tab
    # dates every line of its itinerary.
    run = _run_of_dates(dates)
    out['first_date'], out['last_date'] = (run[0], run[-1]) if run else (None, None)
    # Derive profit_after_vat if the sheet didn't carry it explicitly.
    if out['profit_after_vat'] is None and out['profit_usd'] is not None and out['vat_usd'] is not None:
        out['profit_after_vat'] = round(out['profit_usd'] + out['vat_usd'], 2)
    return code, out


def _fetch_workbook_profit(sheet_id: str, scheduled: bool = True) -> dict:
    url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"
    results: dict = {}
    try:
        resp = requests.get(url, timeout=60)
        resp.raise_for_status()
        wb = load_workbook(io.BytesIO(resp.content), data_only=True, read_only=True)
        for ws in wb.worksheets:
            code, data = _parse_worksheet(ws)
            if code and data and any(v is not None for v in data.values()):
                data['scheduled'] = scheduled
                results[code] = data
        wb.close()
        print(f"[profit_sync] {sheet_id}: parsed profit for {len(results)} tours")
    except Exception as e:
        print(f"[profit_sync] Could not fetch/parse {sheet_id}: {e}")
    return results


def fetch_tour_profit() -> dict:
    """Return {tour_code: {profit_usd, vat_usd, ...}} across all balance files."""
    results: dict = {}
    for key, sheet_id in SHEET_IDS.items():
        results.update(_fetch_workbook_profit(sheet_id, key not in FINANCE_ONLY))
    print(f"[profit_sync] Total: profit for {len(results)} tours")
    return results
