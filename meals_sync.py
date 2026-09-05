"""
Sync restaurant/meal data from the same balance workbooks profit_sync reads —
one tab per tour, each dating its own lunch/dinner lines with a restaurant
name. Live fetch downloads the workbook and reads every tab directly with
openpyxl, the same way profit_sync and archive_sync do; a CSV export only
ever returns one tab, which doesn't fit a one-tab-per-tour workbook, so the
sheets aren't fetched that way here.

If a workbook can't be reached at all, the static fallback data (pre-parsed
by Claude Code, from whenever it was last refreshed) is used for whatever
tours live fetch didn't cover.
"""
import io
import re
import requests
from datetime import date as _date, datetime

from openpyxl import load_workbook

from profit_sync import TOUR_CODE_RE, _norm_code, _num, _usd_from_row, _fix_name

# Same workbooks profit_sync reads (each covers more series than its key
# suggests — KT_DT also has DT1, ZT also has LT). MT_ST is left out: those
# tours have no daily_log/tours rows to attach a meal day to.
SHEET_IDS = {
    "KT_DT": "16NWhGGHR7mXAwRyVH_vmYSrZHx1zxrAR",
    "LN":    "1p5rgt6w_1hGpDr2W3Mug1p7rYWi7L7ZR",
    "ZT":    "1aWUi7GuMFZLuSq1dp2MgP_KV4rmwXGAE",
    "TM":    "1I_mMGVWcel93pNH72fYVM6sYlxOrPj0HHQiSxq2pS_o",
    "HM_HT": "1HCg4JqkNgA9f1SX1gXVr_7mp-WRIu0pDz1wlgrtrgRU",
}

MEAL_LINE_RE = re.compile(r'^(ლანჩი|ვახშამი|ვაშამი)\s*-\s*(.+)$')

# Armenia days never name a restaurant at all — nothing to attach a menu to,
# unlike an own-expense or hotel meal, which at least names a place.
_NO_VENUE_RE = re.compile(r'სომხეთი')


def _parse_workbook_meals(content: bytes) -> dict:
    wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    results: dict = {}
    for ws in wb.worksheets:
        m = TOUR_CODE_RE.search(ws.title or '')
        if not m:
            continue
        code = _norm_code(m.group(1))
        meals = []
        for row in ws.iter_rows(values_only=True):
            if not row or row[0] is None:
                continue
            d = row[0]
            if isinstance(d, datetime):
                d = d.date()
            if not isinstance(d, _date):
                continue
            desc = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ''
            mm = MEAL_LINE_RE.match(desc)
            if not mm:
                continue
            mtype = 'lunch' if mm.group(1) == 'ლანჩი' else 'dinner'
            rest = _fix_name(mm.group(2).strip())
            if not rest or _NO_VENUE_RE.search(rest):
                continue
            cells = list(row)
            gel = _num(cells[3]) if len(cells) > 3 else None
            usd = _usd_from_row(cells)
            meals.append({
                "date": d.isoformat(),
                "meal_type": mtype,
                "restaurant": rest,
                "gel_amount": gel or 0.0,
                "usd_amount": usd or 0.0,
            })
        if meals:
            results[code] = meals
    wb.close()
    return results


def _fetch_sheet(sheet_id: str) -> dict:
    url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"
    try:
        resp = requests.get(url, timeout=60)
        resp.raise_for_status()
        return _parse_workbook_meals(resp.content)
    except Exception as e:
        print(f"[meals_sync] Could not fetch/parse {sheet_id}: {e}")
        return {}


# ── STATIC FALLBACK DATA (pre-parsed from financial sheets) ──────────────────
# Updated via Claude Code when sheets change.

STATIC_MEALS: dict = {
    # ── KT series ────────────────────────────────────────────────────────────
    "KT-0428": [
        {"date":"2026-05-01","meal_type":"degustation","restaurant":"ხარება","gel_amount":346.50,"usd_amount":129.29},
        {"date":"2026-05-01","meal_type":"lunch","restaurant":"ხარება","gel_amount":414.00,"usd_amount":154.48},
        {"date":"2026-05-01","meal_type":"dinner","restaurant":"ახალი აზია (ისანი)","gel_amount":330.00,"usd_amount":123.13},
        {"date":"2026-05-03","meal_type":"dinner","restaurant":"ახალი აზია (ისანი)","gel_amount":330.00,"usd_amount":123.13},
        {"date":"2026-05-04","meal_type":"lunch","restaurant":"სალობიე","gel_amount":311.00,"usd_amount":116.04},
        {"date":"2026-05-04","meal_type":"dinner","restaurant":"კრწანისი","gel_amount":843.20,"usd_amount":314.74},
        {"date":"2026-05-05","meal_type":"lunch","restaurant":"ფასანაური","gel_amount":420.00,"usd_amount":156.72},
        {"date":"2026-05-05","meal_type":"dinner","restaurant":"წანარეთი","gel_amount":354.00,"usd_amount":132.09},
        {"date":"2026-05-06","meal_type":"lunch","restaurant":"ახალი აზია (ისანი)","gel_amount":330.00,"usd_amount":123.13},
    ],
    "KT-0602": [
        {"date":"2026-06-05","meal_type":"degustation","restaurant":"ხარება","gel_amount":346.50,"usd_amount":129.29},
        {"date":"2026-06-05","meal_type":"lunch","restaurant":"ხარება","gel_amount":414.00,"usd_amount":154.48},
        {"date":"2026-06-05","meal_type":"dinner","restaurant":"კტვ ცეკვებით","gel_amount":1105.00,"usd_amount":412.31},
        {"date":"2026-06-07","meal_type":"dinner","restaurant":"ახალი აზია (ისანი)","gel_amount":510.00,"usd_amount":190.30},
        {"date":"2026-06-08","meal_type":"lunch","restaurant":"ფასანაური","gel_amount":311.00,"usd_amount":116.04},
        {"date":"2026-06-08","meal_type":"dinner","restaurant":"მარკო პოლო","gel_amount":843.20,"usd_amount":314.74},
        {"date":"2026-06-09","meal_type":"lunch","restaurant":"სალობიე","gel_amount":420.00,"usd_amount":156.72},
        {"date":"2026-06-09","meal_type":"dinner","restaurant":"დინ შენი","gel_amount":354.00,"usd_amount":132.09},
        {"date":"2026-06-10","meal_type":"dinner","restaurant":"ახალი აზია (ისანი)","gel_amount":0,"usd_amount":0},
    ],
    "KT-0609": [
        {"date":"2026-06-12","meal_type":"degustation","restaurant":"ხარება","gel_amount":346.50,"usd_amount":129.29},
        {"date":"2026-06-12","meal_type":"lunch","restaurant":"ხარება","gel_amount":414.00,"usd_amount":154.48},
        {"date":"2026-06-12","meal_type":"dinner","restaurant":"კტვ","gel_amount":1105.00,"usd_amount":412.31},
        {"date":"2026-06-14","meal_type":"dinner","restaurant":"ახალი აზია (ისანი)","gel_amount":330.00,"usd_amount":123.13},
        {"date":"2026-06-15","meal_type":"lunch","restaurant":"ფასანაური","gel_amount":311.00,"usd_amount":116.04},
        {"date":"2026-06-15","meal_type":"dinner","restaurant":"მარკო პოლო","gel_amount":843.20,"usd_amount":314.74},
        {"date":"2026-06-16","meal_type":"lunch","restaurant":"სალობიე","gel_amount":420.00,"usd_amount":156.72},
        {"date":"2026-06-16","meal_type":"dinner","restaurant":"დინ შენი","gel_amount":354.00,"usd_amount":132.09},
        {"date":"2026-06-17","meal_type":"dinner","restaurant":"ახალი აზია (ისანი)","gel_amount":0,"usd_amount":0},
    ],
    # ── DT1 series ───────────────────────────────────────────────────────────
    "DT1-0524": [
        {"date":"2026-05-26","meal_type":"degustation","restaurant":"ხარება","gel_amount":657.90,"usd_amount":245.58},
        {"date":"2026-05-26","meal_type":"lunch","restaurant":"ხარება","gel_amount":598.50,"usd_amount":223.40},
        {"date":"2026-05-26","meal_type":"dinner","restaurant":"კტვ","gel_amount":1235.00,"usd_amount":460.82},
        {"date":"2026-05-28","meal_type":"dinner","restaurant":"ახალი აზია (ისანი)","gel_amount":570.00,"usd_amount":212.69},
        {"date":"2026-05-29","meal_type":"lunch","restaurant":"სალობიე","gel_amount":493.00,"usd_amount":183.96},
        {"date":"2026-05-29","meal_type":"dinner","restaurant":"დინ შენი","gel_amount":665.00,"usd_amount":248.13},
        {"date":"2026-05-30","meal_type":"lunch","restaurant":"ფასანაური","gel_amount":512.00,"usd_amount":191.04},
        {"date":"2026-05-30","meal_type":"dinner","restaurant":"ოქროს საწმისი","gel_amount":532.00,"usd_amount":198.51},
        {"date":"2026-05-31","meal_type":"lunch","restaurant":"ახალი აზია (ისანი)","gel_amount":570.00,"usd_amount":212.69},
    ],
    "DT1-0531": [
        {"date":"2026-06-02","meal_type":"degustation","restaurant":"ხარება","gel_amount":504.00,"usd_amount":188.13},
        {"date":"2026-06-02","meal_type":"lunch","restaurant":"ხარება","gel_amount":534.15,"usd_amount":199.38},
        {"date":"2026-06-02","meal_type":"dinner","restaurant":"კტვ ცეკვებით","gel_amount":1040.00,"usd_amount":388.06},
        {"date":"2026-06-04","meal_type":"dinner","restaurant":"ახალი აზია (ისანი)","gel_amount":480.00,"usd_amount":179.10},
        {"date":"2026-06-05","meal_type":"lunch","restaurant":"სალობიე","gel_amount":459.00,"usd_amount":171.27},
        {"date":"2026-06-05","meal_type":"dinner","restaurant":"დინ შენი","gel_amount":560.00,"usd_amount":208.96},
        {"date":"2026-06-06","meal_type":"lunch","restaurant":"ფასანაური","gel_amount":420.00,"usd_amount":156.72},
        {"date":"2026-06-06","meal_type":"dinner","restaurant":"ოქროს საწმისი","gel_amount":354.00,"usd_amount":132.09},
        {"date":"2026-06-07","meal_type":"lunch","restaurant":"ახალი აზია (ისანი)","gel_amount":480.00,"usd_amount":179.10},
    ],
    # ── DT2 series ───────────────────────────────────────────────────────────
    "DT2-0511": [
        {"date":"2026-05-13","meal_type":"degustation","restaurant":"ხარება","gel_amount":657.90,"usd_amount":245.58},
        {"date":"2026-05-13","meal_type":"lunch","restaurant":"ხარება","gel_amount":598.50,"usd_amount":223.40},
        {"date":"2026-05-13","meal_type":"dinner","restaurant":"კტვ","gel_amount":1235.00,"usd_amount":460.82},
        {"date":"2026-05-15","meal_type":"dinner","restaurant":"ახალი აზია (ისანი)","gel_amount":570.00,"usd_amount":212.69},
        {"date":"2026-05-16","meal_type":"lunch","restaurant":"სალობიე","gel_amount":487.00,"usd_amount":181.72},
        {"date":"2026-05-16","meal_type":"dinner","restaurant":"დინ შენი","gel_amount":665.00,"usd_amount":248.13},
        {"date":"2026-05-17","meal_type":"lunch","restaurant":"ფასანაური","gel_amount":550.00,"usd_amount":205.22},
        {"date":"2026-05-17","meal_type":"dinner","restaurant":"ოქროს საწმისი","gel_amount":532.00,"usd_amount":198.51},
        {"date":"2026-05-18","meal_type":"lunch","restaurant":"ახალი აზია (ისანი)","gel_amount":510.00,"usd_amount":190.30},
    ],
    "DT2-0524": [
        {"date":"2026-05-26","meal_type":"degustation","restaurant":"ხარება","gel_amount":657.90,"usd_amount":245.58},
        {"date":"2026-05-26","meal_type":"lunch","restaurant":"ხარება","gel_amount":598.50,"usd_amount":223.40},
        {"date":"2026-05-26","meal_type":"dinner","restaurant":"კტვ ცეკვებით","gel_amount":1235.00,"usd_amount":460.82},
        {"date":"2026-05-28","meal_type":"dinner","restaurant":"ახალი აზია (ისანი)","gel_amount":570.00,"usd_amount":212.69},
        {"date":"2026-05-29","meal_type":"lunch","restaurant":"გურამიშვილის მარანი","gel_amount":1843.00,"usd_amount":687.69},
        {"date":"2026-05-29","meal_type":"dinner","restaurant":"დინ შენი","gel_amount":665.00,"usd_amount":248.13},
        {"date":"2026-05-30","meal_type":"lunch","restaurant":"ფასანაური","gel_amount":512.00,"usd_amount":191.04},
        {"date":"2026-05-30","meal_type":"dinner","restaurant":"ოქროს საწმისი","gel_amount":532.00,"usd_amount":198.51},
        {"date":"2026-05-31","meal_type":"dinner","restaurant":"ახალი აზია (ისანი)","gel_amount":570.00,"usd_amount":212.69},
    ],
    "DT2-0601": [
        {"date":"2026-06-03","meal_type":"degustation","restaurant":"ხარება","gel_amount":315.00,"usd_amount":117.58},
        {"date":"2026-06-03","meal_type":"lunch","restaurant":"ხარება","gel_amount":398.25,"usd_amount":148.66},
        {"date":"2026-06-03","meal_type":"dinner","restaurant":"კტვ","gel_amount":650.00,"usd_amount":242.54},
        {"date":"2026-06-05","meal_type":"dinner","restaurant":"ახალი აზია (ისანი)","gel_amount":300.00,"usd_amount":111.94},
        {"date":"2026-06-06","meal_type":"lunch","restaurant":"ფასანაური","gel_amount":487.00,"usd_amount":181.72},
        {"date":"2026-06-06","meal_type":"dinner","restaurant":"ოქროს საწმისი","gel_amount":665.00,"usd_amount":248.13},
        {"date":"2026-06-07","meal_type":"lunch","restaurant":"ახალი აზია (ისანი)","gel_amount":300.00,"usd_amount":111.94},
    ],
    # ── LN series ────────────────────────────────────────────────────────────
    "LN-0501": [
        {"date":"2026-05-04","meal_type":"degustation","restaurant":"ხარება","gel_amount":567.00,"usd_amount":211.57},
        {"date":"2026-05-04","meal_type":"lunch","restaurant":"ხარება","gel_amount":642.15,"usd_amount":239.61},
        {"date":"2026-05-04","meal_type":"dinner","restaurant":"ახალი აზია (ისანი)","gel_amount":540.00,"usd_amount":201.49},
        {"date":"2026-05-06","meal_type":"dinner","restaurant":"ახალციხე ინნ","gel_amount":0,"usd_amount":0},
        {"date":"2026-05-07","meal_type":"lunch","restaurant":"ზღაპარი","gel_amount":385.00,"usd_amount":143.66},
        {"date":"2026-05-08","meal_type":"lunch","restaurant":"მარტვილი","gel_amount":531.00,"usd_amount":198.13},
        {"date":"2026-05-08","meal_type":"dinner","restaurant":"გორი ინნ","gel_amount":0,"usd_amount":0},
        {"date":"2026-05-09","meal_type":"lunch","restaurant":"ფასანაური","gel_amount":721.00,"usd_amount":269.03},
        {"date":"2026-05-09","meal_type":"dinner","restaurant":"მარკო პოლო","gel_amount":0,"usd_amount":0},
        {"date":"2026-05-10","meal_type":"lunch","restaurant":"სალობიე","gel_amount":473.00,"usd_amount":176.49},
        {"date":"2026-05-10","meal_type":"dinner","restaurant":"კტვ","gel_amount":1170.00,"usd_amount":436.57},
        {"date":"2026-05-11","meal_type":"lunch","restaurant":"ახალი აზია (ისანი)","gel_amount":540.00,"usd_amount":201.57},
    ],
    "LN-0508": [
        {"date":"2026-05-11","meal_type":"degustation","restaurant":"ხარება","gel_amount":630.00,"usd_amount":235.07},
        {"date":"2026-05-11","meal_type":"lunch","restaurant":"ხარება","gel_amount":664.65,"usd_amount":248.00},
        {"date":"2026-05-11","meal_type":"dinner","restaurant":"კტვ","gel_amount":1300.00,"usd_amount":485.07},
        {"date":"2026-05-13","meal_type":"dinner","restaurant":"ახალციხე ინნ","gel_amount":0,"usd_amount":0},
        {"date":"2026-05-14","meal_type":"lunch","restaurant":"ზღაპარი","gel_amount":452.00,"usd_amount":168.66},
        {"date":"2026-05-15","meal_type":"lunch","restaurant":"მარტვილი","gel_amount":572.00,"usd_amount":213.43},
        {"date":"2026-05-15","meal_type":"dinner","restaurant":"გორი ინნ","gel_amount":0,"usd_amount":0},
        {"date":"2026-05-16","meal_type":"lunch","restaurant":"ფასანაური","gel_amount":640.00,"usd_amount":238.81},
        {"date":"2026-05-16","meal_type":"dinner","restaurant":"მარკო პოლო","gel_amount":0,"usd_amount":0},
        {"date":"2026-05-17","meal_type":"lunch","restaurant":"სალობიე","gel_amount":499.00,"usd_amount":186.19},
        {"date":"2026-05-17","meal_type":"dinner","restaurant":"დინ შენი","gel_amount":700.00,"usd_amount":261.19},
        {"date":"2026-05-18","meal_type":"dinner","restaurant":"ახალი აზია (ისანი)","gel_amount":570.00,"usd_amount":212.69},
    ],
    "LN-0518": [
        {"date":"2026-05-21","meal_type":"degustation","restaurant":"ხარება","gel_amount":598.50,"usd_amount":223.32},
        {"date":"2026-05-21","meal_type":"lunch","restaurant":"ხარება","gel_amount":630.00,"usd_amount":235.07},
        {"date":"2026-05-21","meal_type":"dinner","restaurant":"კტვ","gel_amount":1235.00,"usd_amount":460.82},
        {"date":"2026-05-23","meal_type":"dinner","restaurant":"ახალციხე ინნ","gel_amount":0,"usd_amount":0},
        {"date":"2026-05-24","meal_type":"lunch","restaurant":"ზღაპარი","gel_amount":390.00,"usd_amount":145.52},
        {"date":"2026-05-25","meal_type":"lunch","restaurant":"მარტვილი","gel_amount":571.00,"usd_amount":213.06},
        {"date":"2026-05-25","meal_type":"dinner","restaurant":"გორი ინნ","gel_amount":0,"usd_amount":0},
        {"date":"2026-05-26","meal_type":"lunch","restaurant":"ფასანაური","gel_amount":555.00,"usd_amount":207.09},
        {"date":"2026-05-26","meal_type":"dinner","restaurant":"მარკო პოლო","gel_amount":0,"usd_amount":0},
        {"date":"2026-05-27","meal_type":"lunch","restaurant":"სალობიე","gel_amount":493.00,"usd_amount":183.96},
        {"date":"2026-05-27","meal_type":"dinner","restaurant":"დინ შენი","gel_amount":665.00,"usd_amount":248.13},
        {"date":"2026-05-28","meal_type":"dinner","restaurant":"ახალი აზია (ისანი)","gel_amount":570.00,"usd_amount":212.69},
    ],
    "LN-0525": [
        {"date":"2026-05-28","meal_type":"degustation","restaurant":"ხარება","gel_amount":598.50,"usd_amount":223.32},
        {"date":"2026-05-28","meal_type":"lunch","restaurant":"ხარება","gel_amount":664.65,"usd_amount":248.00},
        {"date":"2026-05-28","meal_type":"dinner","restaurant":"კტვ ცეკვებით","gel_amount":1300.00,"usd_amount":485.07},
        {"date":"2026-05-30","meal_type":"dinner","restaurant":"ახალციხე ინნ","gel_amount":0,"usd_amount":0},
        {"date":"2026-05-31","meal_type":"lunch","restaurant":"ზღაპარი","gel_amount":440.00,"usd_amount":164.18},
        {"date":"2026-06-01","meal_type":"lunch","restaurant":"მარტვილი","gel_amount":546.00,"usd_amount":203.73},
        {"date":"2026-06-01","meal_type":"dinner","restaurant":"გორი ინნ","gel_amount":0,"usd_amount":0},
        {"date":"2026-06-02","meal_type":"lunch","restaurant":"ფასანაური","gel_amount":645.00,"usd_amount":240.67},
        {"date":"2026-06-02","meal_type":"dinner","restaurant":"მარკო პოლო","gel_amount":0,"usd_amount":0},
        {"date":"2026-06-03","meal_type":"lunch","restaurant":"სალობიე","gel_amount":544.00,"usd_amount":202.99},
        {"date":"2026-06-03","meal_type":"dinner","restaurant":"დინ შენი","gel_amount":700.00,"usd_amount":261.19},
        {"date":"2026-06-04","meal_type":"dinner","restaurant":"ახალი აზია (ისანი)","gel_amount":600.00,"usd_amount":223.88},
    ],
    "LN-0527": [
        {"date":"2026-05-30","meal_type":"degustation","restaurant":"ხარება","gel_amount":598.50,"usd_amount":223.32},
        {"date":"2026-05-30","meal_type":"lunch","restaurant":"ხარება","gel_amount":657.90,"usd_amount":245.49},
        {"date":"2026-05-30","meal_type":"dinner","restaurant":"კტვ","gel_amount":1235.00,"usd_amount":460.82},
        {"date":"2026-06-01","meal_type":"dinner","restaurant":"ახალციხე ინნ","gel_amount":0,"usd_amount":0},
        {"date":"2026-06-02","meal_type":"lunch","restaurant":"ზღაპარი","gel_amount":390.00,"usd_amount":145.52},
        {"date":"2026-06-03","meal_type":"lunch","restaurant":"მარტვილი","gel_amount":542.00,"usd_amount":202.24},
        {"date":"2026-06-03","meal_type":"dinner","restaurant":"გორი ინნ","gel_amount":0,"usd_amount":0},
        {"date":"2026-06-04","meal_type":"lunch","restaurant":"ფასანაური","gel_amount":635.00,"usd_amount":236.94},
        {"date":"2026-06-04","meal_type":"dinner","restaurant":"მარკო პოლო","gel_amount":0,"usd_amount":0},
        {"date":"2026-06-05","meal_type":"lunch","restaurant":"სალობიე","gel_amount":538.00,"usd_amount":200.75},
        {"date":"2026-06-05","meal_type":"dinner","restaurant":"დინ შენი","gel_amount":665.00,"usd_amount":248.13},
        {"date":"2026-06-06","meal_type":"lunch","restaurant":"ახალი აზია (ისანი)","gel_amount":570.00,"usd_amount":212.69},
    ],
    "LN-0601": [
        {"date":"2026-06-04","meal_type":"degustation","restaurant":"ხარება","gel_amount":567.00,"usd_amount":211.57},
        {"date":"2026-06-04","meal_type":"lunch","restaurant":"ხარება","gel_amount":663.75,"usd_amount":247.67},
        {"date":"2026-06-04","meal_type":"dinner","restaurant":"კტვ","gel_amount":1170.00,"usd_amount":436.57},
        {"date":"2026-06-06","meal_type":"dinner","restaurant":"ახალციხე ინნ","gel_amount":0,"usd_amount":0},
        {"date":"2026-06-07","meal_type":"lunch","restaurant":"ზღაპარი","gel_amount":385.00,"usd_amount":143.66},
        {"date":"2026-06-08","meal_type":"lunch","restaurant":"მარტვილი","gel_amount":537.00,"usd_amount":200.37},
        {"date":"2026-06-08","meal_type":"dinner","restaurant":"გორი ინნ","gel_amount":0,"usd_amount":0},
        {"date":"2026-06-09","meal_type":"lunch","restaurant":"ფასანაური","gel_amount":329.00,"usd_amount":122.76},
        {"date":"2026-06-09","meal_type":"dinner","restaurant":"მარკო პოლო","gel_amount":0,"usd_amount":0},
        {"date":"2026-06-10","meal_type":"lunch","restaurant":"სალობიე","gel_amount":410.00,"usd_amount":152.99},
        {"date":"2026-06-10","meal_type":"dinner","restaurant":"დინ შენი","gel_amount":665.00,"usd_amount":248.13},
        {"date":"2026-06-11","meal_type":"dinner","restaurant":"ახალი აზია (ისანი)","gel_amount":560.93,"usd_amount":209.30},
    ],
    "LN-0603": [
        {"date":"2026-06-06","meal_type":"degustation","restaurant":"ხარება","gel_amount":598.50,"usd_amount":223.32},
        {"date":"2026-06-06","meal_type":"lunch","restaurant":"ხარება","gel_amount":630.00,"usd_amount":235.07},
        {"date":"2026-06-06","meal_type":"dinner","restaurant":"კტვ","gel_amount":1040.00,"usd_amount":388.06},
        {"date":"2026-06-08","meal_type":"dinner","restaurant":"ახალციხე ინნ","gel_amount":0,"usd_amount":0},
        {"date":"2026-06-09","meal_type":"lunch","restaurant":"ზღაპარი","gel_amount":390.00,"usd_amount":145.52},
        {"date":"2026-06-10","meal_type":"lunch","restaurant":"მარტვილი","gel_amount":330.00,"usd_amount":123.13},
        {"date":"2026-06-10","meal_type":"dinner","restaurant":"გორი ინნ","gel_amount":0,"usd_amount":0},
        {"date":"2026-06-11","meal_type":"lunch","restaurant":"ფასანაური","gel_amount":329.00,"usd_amount":122.76},
        {"date":"2026-06-11","meal_type":"dinner","restaurant":"მარკო პოლო","gel_amount":0,"usd_amount":0},
        {"date":"2026-06-12","meal_type":"lunch","restaurant":"სალობიე","gel_amount":410.00,"usd_amount":152.99},
        {"date":"2026-06-12","meal_type":"dinner","restaurant":"დინ შენი","gel_amount":665.00,"usd_amount":248.13},
        {"date":"2026-06-13","meal_type":"dinner","restaurant":"ახალი აზია (ისანი)","gel_amount":560.93,"usd_amount":209.30},
    ],
    "LN-0608": [
        {"date":"2026-06-11","meal_type":"degustation","restaurant":"ხარება","gel_amount":598.50,"usd_amount":223.32},
        {"date":"2026-06-11","meal_type":"lunch","restaurant":"ხარება","gel_amount":630.00,"usd_amount":235.07},
        {"date":"2026-06-11","meal_type":"dinner","restaurant":"კტვ","gel_amount":1040.00,"usd_amount":388.06},
        {"date":"2026-06-13","meal_type":"dinner","restaurant":"ახალციხე ინნ","gel_amount":0,"usd_amount":0},
        {"date":"2026-06-14","meal_type":"lunch","restaurant":"ზღაპარი","gel_amount":390.00,"usd_amount":145.52},
        {"date":"2026-06-15","meal_type":"lunch","restaurant":"მარტვილი","gel_amount":330.00,"usd_amount":123.13},
        {"date":"2026-06-15","meal_type":"dinner","restaurant":"გორი ინნ","gel_amount":0,"usd_amount":0},
        {"date":"2026-06-16","meal_type":"lunch","restaurant":"ფასანაური","gel_amount":329.00,"usd_amount":122.76},
        {"date":"2026-06-16","meal_type":"dinner","restaurant":"მარკო პოლო","gel_amount":0,"usd_amount":0},
        {"date":"2026-06-17","meal_type":"lunch","restaurant":"სალობიე","gel_amount":410.00,"usd_amount":152.99},
        {"date":"2026-06-17","meal_type":"dinner","restaurant":"დინ შენი","gel_amount":665.00,"usd_amount":248.13},
        {"date":"2026-06-18","meal_type":"dinner","restaurant":"ახალი აზია (ისანი)","gel_amount":560.93,"usd_amount":209.30},
    ],
    # ── ZT series ────────────────────────────────────────────────────────────
    "ZT-0427": [
        {"date":"2026-04-30","meal_type":"degustation","restaurant":"ხარება","gel_amount":598.50,"usd_amount":223.32},
        {"date":"2026-04-30","meal_type":"lunch","restaurant":"ხარება","gel_amount":657.90,"usd_amount":245.49},
        {"date":"2026-04-30","meal_type":"dinner","restaurant":"ახალი აზია (ისანი)","gel_amount":570.00,"usd_amount":212.77},
        {"date":"2026-05-02","meal_type":"dinner","restaurant":"ახალციხე ინნ","gel_amount":0,"usd_amount":0},
        {"date":"2026-05-03","meal_type":"lunch","restaurant":"ზღაპარი","gel_amount":476.00,"usd_amount":177.61},
        {"date":"2026-05-04","meal_type":"lunch","restaurant":"დიარონი","gel_amount":539.84,"usd_amount":201.43},
        {"date":"2026-05-04","meal_type":"dinner","restaurant":"ლუშნუ ქორი","gel_amount":497.00,"usd_amount":185.52},
        {"date":"2026-05-05","meal_type":"lunch","restaurant":"ენგური","gel_amount":513.45,"usd_amount":191.59},
        {"date":"2026-05-05","meal_type":"dinner","restaurant":"ლუიზასთან","gel_amount":500.00,"usd_amount":186.57},
        {"date":"2026-05-06","meal_type":"lunch","restaurant":"მარტვილი","gel_amount":540.00,"usd_amount":201.49},
        {"date":"2026-05-06","meal_type":"dinner","restaurant":"გორი ინნ","gel_amount":0,"usd_amount":0},
        {"date":"2026-05-07","meal_type":"lunch","restaurant":"სალობიე","gel_amount":487.00,"usd_amount":181.72},
        {"date":"2026-05-07","meal_type":"dinner","restaurant":"მარკო პოლო","gel_amount":0,"usd_amount":0},
        {"date":"2026-05-08","meal_type":"lunch","restaurant":"ახალი აზია (ისანი)","gel_amount":570.00,"usd_amount":212.69},
        {"date":"2026-05-08","meal_type":"dinner","restaurant":"კტვ","gel_amount":1235.00,"usd_amount":460.82},
    ],
    "ZT-0504": [
        {"date":"2026-05-07","meal_type":"degustation","restaurant":"ხარება","gel_amount":598.50,"usd_amount":223.32},
        {"date":"2026-05-07","meal_type":"lunch","restaurant":"ხარება","gel_amount":657.90,"usd_amount":245.49},
        {"date":"2026-05-07","meal_type":"dinner","restaurant":"კტვ","gel_amount":1297.00,"usd_amount":483.96},
        {"date":"2026-05-09","meal_type":"dinner","restaurant":"ახალციხე ინნ","gel_amount":0,"usd_amount":0},
        {"date":"2026-05-10","meal_type":"lunch","restaurant":"ზღაპარი","gel_amount":432.00,"usd_amount":161.19},
        {"date":"2026-05-11","meal_type":"lunch","restaurant":"დიარონი","gel_amount":537.60,"usd_amount":200.60},
        {"date":"2026-05-11","meal_type":"dinner","restaurant":"ლუშნუ ქორი","gel_amount":586.00,"usd_amount":218.74},
        {"date":"2026-05-12","meal_type":"lunch","restaurant":"ენგური","gel_amount":494.50,"usd_amount":184.51},
        {"date":"2026-05-12","meal_type":"dinner","restaurant":"ლუიზასთან","gel_amount":531.00,"usd_amount":198.21},
        {"date":"2026-05-13","meal_type":"lunch","restaurant":"კვამლი","gel_amount":579.70,"usd_amount":216.31},
        {"date":"2026-05-13","meal_type":"dinner","restaurant":"გორი ინნ","gel_amount":0,"usd_amount":0},
        {"date":"2026-05-14","meal_type":"lunch","restaurant":"ფასანაური","gel_amount":694.00,"usd_amount":258.96},
        {"date":"2026-05-14","meal_type":"dinner","restaurant":"მარკო პოლო","gel_amount":0,"usd_amount":0},
        {"date":"2026-05-15","meal_type":"lunch","restaurant":"დინ შენი","gel_amount":665.00,"usd_amount":248.23},
        {"date":"2026-05-15","meal_type":"dinner","restaurant":"ახალი აზია ისანი","gel_amount":570.00,"usd_amount":212.69},
    ],
    "ZT-0511": [
        {"date":"2026-05-14","meal_type":"degustation","restaurant":"ხარება","gel_amount":567.00,"usd_amount":211.57},
        {"date":"2026-05-14","meal_type":"lunch","restaurant":"ხარება","gel_amount":642.15,"usd_amount":239.61},
        {"date":"2026-05-14","meal_type":"dinner","restaurant":"კტვ","gel_amount":1170.00,"usd_amount":436.57},
        {"date":"2026-05-17","meal_type":"lunch","restaurant":"ახალი აზია ისანი","gel_amount":476.00,"usd_amount":177.61},
        {"date":"2026-05-17","meal_type":"dinner","restaurant":"დინ შენი","gel_amount":630.00,"usd_amount":235.07},
        {"date":"2026-05-18","meal_type":"lunch","restaurant":"სალობიე","gel_amount":481.00,"usd_amount":179.48},
        {"date":"2026-05-18","meal_type":"dinner","restaurant":"მარკო პოლო","gel_amount":0,"usd_amount":0},
        {"date":"2026-05-19","meal_type":"lunch","restaurant":"ცენტრალ პაბ გორი","gel_amount":573.00,"usd_amount":213.81},
        {"date":"2026-05-19","meal_type":"dinner","restaurant":"ახალციხე ინნ","gel_amount":0,"usd_amount":0},
        {"date":"2026-05-20","meal_type":"lunch","restaurant":"კვამლი","gel_amount":469.70,"usd_amount":175.26},
        {"date":"2026-05-20","meal_type":"dinner","restaurant":"ბერიძეები","gel_amount":410.00,"usd_amount":152.99},
        {"date":"2026-05-21","meal_type":"lunch","restaurant":"ბერიძეები","gel_amount":410.00,"usd_amount":152.99},
        {"date":"2026-05-21","meal_type":"dinner","restaurant":"ოქროს საწმისი მცხეთა","gel_amount":504.00,"usd_amount":188.06},
        {"date":"2026-05-22","meal_type":"lunch","restaurant":"ახალი აზია ისანი","gel_amount":625.00,"usd_amount":233.30},
    ],
    "ZT-0518": [
        {"date":"2026-05-21","meal_type":"degustation","restaurant":"ხარება","gel_amount":657.90,"usd_amount":248.26},
        {"date":"2026-05-21","meal_type":"lunch","restaurant":"ხარება","gel_amount":664.65,"usd_amount":250.81},
        {"date":"2026-05-21","meal_type":"dinner","restaurant":"კტვ ცეკვებით","gel_amount":1300.00,"usd_amount":490.57},
        {"date":"2026-05-23","meal_type":"dinner","restaurant":"ახალციხე ინნ","gel_amount":0,"usd_amount":0},
        {"date":"2026-05-24","meal_type":"lunch","restaurant":"ზღაპარი","gel_amount":396.00,"usd_amount":149.43},
        {"date":"2026-05-25","meal_type":"lunch","restaurant":"დიარონი","gel_amount":532.00,"usd_amount":200.75},
        {"date":"2026-05-25","meal_type":"dinner","restaurant":"ლუშნუ ქორი","gel_amount":603.00,"usd_amount":227.55},
        {"date":"2026-05-26","meal_type":"lunch","restaurant":"ენგური","gel_amount":503.00,"usd_amount":189.81},
        {"date":"2026-05-26","meal_type":"dinner","restaurant":"ლუიზასთან","gel_amount":499.00,"usd_amount":186.26},
        {"date":"2026-05-27","meal_type":"lunch","restaurant":"დიარონი","gel_amount":543.20,"usd_amount":204.98},
        {"date":"2026-05-27","meal_type":"dinner","restaurant":"გორი ინნ","gel_amount":0,"usd_amount":0},
        {"date":"2026-05-28","meal_type":"lunch","restaurant":"სალობიე","gel_amount":493.00,"usd_amount":186.04},
        {"date":"2026-05-28","meal_type":"dinner","restaurant":"გუდაური ინნ","gel_amount":0,"usd_amount":0},
        {"date":"2026-05-29","meal_type":"lunch","restaurant":"დინ შენი","gel_amount":700.00,"usd_amount":264.15},
        {"date":"2026-05-29","meal_type":"dinner","restaurant":"ახალი აზია ისანი","gel_amount":600.00,"usd_amount":226.42},
    ],
    "ZT-0525": [
        {"date":"2026-05-28","meal_type":"degustation","restaurant":"ხარება","gel_amount":535.50,"usd_amount":199.81},
        {"date":"2026-05-28","meal_type":"lunch","restaurant":"ხარება","gel_amount":657.90,"usd_amount":245.49},
        {"date":"2026-05-28","meal_type":"dinner","restaurant":"კტვ","gel_amount":1235.00,"usd_amount":460.82},
        {"date":"2026-05-30","meal_type":"dinner","restaurant":"ახალციხე ინნ","gel_amount":0,"usd_amount":0},
        {"date":"2026-05-31","meal_type":"lunch","restaurant":"ზღაპარი","gel_amount":396.00,"usd_amount":147.76},
        {"date":"2026-06-01","meal_type":"lunch","restaurant":"დიარონი","gel_amount":532.00,"usd_amount":198.51},
        {"date":"2026-06-01","meal_type":"dinner","restaurant":"ლუშნუ ქორი","gel_amount":586.00,"usd_amount":218.74},
        {"date":"2026-06-02","meal_type":"lunch","restaurant":"ენგური","gel_amount":504.00,"usd_amount":188.06},
        {"date":"2026-06-02","meal_type":"dinner","restaurant":"ლუიზასთან","gel_amount":500.00,"usd_amount":186.64},
        {"date":"2026-06-03","meal_type":"lunch","restaurant":"დიარონი","gel_amount":540.00,"usd_amount":201.49},
        {"date":"2026-06-03","meal_type":"dinner","restaurant":"გორი ინნ","gel_amount":0,"usd_amount":0},
        {"date":"2026-06-04","meal_type":"lunch","restaurant":"სალობიე","gel_amount":538.00,"usd_amount":200.75},
        {"date":"2026-06-04","meal_type":"dinner","restaurant":"მარკო პოლო","gel_amount":0,"usd_amount":0},
        {"date":"2026-06-05","meal_type":"lunch","restaurant":"დინ შენი","gel_amount":665.00,"usd_amount":248.23},
        {"date":"2026-06-05","meal_type":"dinner","restaurant":"ახალი აზია ისანი","gel_amount":570.00,"usd_amount":212.69},
    ],
    "ZT-0601": [
        {"date":"2026-06-04","meal_type":"degustation","restaurant":"ხარება","gel_amount":598.50,"usd_amount":223.32},
        {"date":"2026-06-04","meal_type":"lunch","restaurant":"ხარება","gel_amount":670.50,"usd_amount":250.19},
        {"date":"2026-06-04","meal_type":"dinner","restaurant":"კტვ","gel_amount":1235.00,"usd_amount":460.82},
        {"date":"2026-06-06","meal_type":"dinner","restaurant":"ახალციხე ინნ","gel_amount":0,"usd_amount":0},
        {"date":"2026-06-07","meal_type":"lunch","restaurant":"ზღაპარი","gel_amount":390.00,"usd_amount":145.52},
        {"date":"2026-06-08","meal_type":"lunch","restaurant":"დიარონი","gel_amount":532.00,"usd_amount":198.51},
        {"date":"2026-06-08","meal_type":"dinner","restaurant":"ლუშნუ ქორი","gel_amount":586.00,"usd_amount":218.74},
        {"date":"2026-06-09","meal_type":"lunch","restaurant":"ენგური","gel_amount":330.00,"usd_amount":123.13},
        {"date":"2026-06-09","meal_type":"dinner","restaurant":"ლუიზასთან","gel_amount":499.00,"usd_amount":186.26},
        {"date":"2026-06-10","meal_type":"lunch","restaurant":"დიარონი","gel_amount":540.00,"usd_amount":201.49},
        {"date":"2026-06-10","meal_type":"dinner","restaurant":"გორი ინნ","gel_amount":0,"usd_amount":0},
        {"date":"2026-06-11","meal_type":"lunch","restaurant":"სალობიე","gel_amount":410.00,"usd_amount":152.99},
        {"date":"2026-06-11","meal_type":"dinner","restaurant":"მარკო პოლო","gel_amount":0,"usd_amount":0},
        {"date":"2026-06-12","meal_type":"lunch","restaurant":"დინ შენი","gel_amount":700.00,"usd_amount":261.29},
        {"date":"2026-06-12","meal_type":"dinner","restaurant":"ახალი აზია ისანი","gel_amount":560.93,"usd_amount":209.30},
    ],
    "ZT-0608": [
        {"date":"2026-06-11","meal_type":"degustation","restaurant":"ხარება","gel_amount":598.50,"usd_amount":223.32},
        {"date":"2026-06-11","meal_type":"lunch","restaurant":"ხარება","gel_amount":670.50,"usd_amount":250.19},
        {"date":"2026-06-11","meal_type":"dinner","restaurant":"კტვ","gel_amount":1235.00,"usd_amount":460.82},
        {"date":"2026-06-13","meal_type":"dinner","restaurant":"ახალციხე ინნ","gel_amount":0,"usd_amount":0},
        {"date":"2026-06-14","meal_type":"lunch","restaurant":"ზღაპარი","gel_amount":396.00,"usd_amount":147.76},
        {"date":"2026-06-15","meal_type":"lunch","restaurant":"დიარონი","gel_amount":532.00,"usd_amount":198.51},
        {"date":"2026-06-15","meal_type":"dinner","restaurant":"ლუშნუ ქორი","gel_amount":586.00,"usd_amount":218.74},
        {"date":"2026-06-16","meal_type":"lunch","restaurant":"ენგური","gel_amount":330.00,"usd_amount":123.13},
        {"date":"2026-06-16","meal_type":"dinner","restaurant":"ლუიზასთან","gel_amount":499.00,"usd_amount":186.26},
        {"date":"2026-06-17","meal_type":"lunch","restaurant":"დიარონი","gel_amount":540.00,"usd_amount":201.49},
        {"date":"2026-06-17","meal_type":"dinner","restaurant":"გორი ინნ","gel_amount":0,"usd_amount":0},
        {"date":"2026-06-18","meal_type":"lunch","restaurant":"სალობიე","gel_amount":410.00,"usd_amount":152.99},
        {"date":"2026-06-18","meal_type":"dinner","restaurant":"მარკო პოლო","gel_amount":0,"usd_amount":0},
        {"date":"2026-06-19","meal_type":"lunch","restaurant":"დინ შენი","gel_amount":700.00,"usd_amount":261.29},
        {"date":"2026-06-19","meal_type":"dinner","restaurant":"ახალი აზია ისანი","gel_amount":560.93,"usd_amount":209.30},
    ],
    "ZT-0615": [
        {"date":"2026-06-18","meal_type":"degustation","restaurant":"ხარება","gel_amount":598.50,"usd_amount":223.32},
        {"date":"2026-06-18","meal_type":"lunch","restaurant":"ხარება","gel_amount":670.50,"usd_amount":250.19},
        {"date":"2026-06-18","meal_type":"dinner","restaurant":"კტვ","gel_amount":1235.00,"usd_amount":460.82},
        {"date":"2026-06-20","meal_type":"dinner","restaurant":"ახალციხე ინნ","gel_amount":0,"usd_amount":0},
        {"date":"2026-06-21","meal_type":"lunch","restaurant":"ზღაპარი","gel_amount":396.00,"usd_amount":147.76},
        {"date":"2026-06-22","meal_type":"lunch","restaurant":"დიარონი","gel_amount":532.00,"usd_amount":198.51},
        {"date":"2026-06-22","meal_type":"dinner","restaurant":"ლუშნუ ქორი","gel_amount":586.00,"usd_amount":218.74},
        {"date":"2026-06-23","meal_type":"lunch","restaurant":"ენგური","gel_amount":330.00,"usd_amount":123.13},
        {"date":"2026-06-23","meal_type":"dinner","restaurant":"ლუიზასთან","gel_amount":499.00,"usd_amount":186.26},
        {"date":"2026-06-24","meal_type":"lunch","restaurant":"დიარონი","gel_amount":540.00,"usd_amount":201.49},
        {"date":"2026-06-24","meal_type":"dinner","restaurant":"გორი ინნ","gel_amount":0,"usd_amount":0},
        {"date":"2026-06-25","meal_type":"lunch","restaurant":"სალობიე","gel_amount":410.00,"usd_amount":152.99},
        {"date":"2026-06-25","meal_type":"dinner","restaurant":"მარკო პოლო","gel_amount":0,"usd_amount":0},
        {"date":"2026-06-26","meal_type":"lunch","restaurant":"დინ შენი","gel_amount":700.00,"usd_amount":261.29},
        {"date":"2026-06-26","meal_type":"dinner","restaurant":"ახალი აზია ისანი","gel_amount":560.93,"usd_amount":209.30},
    ],
}


def fetch_all_meals() -> dict:
    """
    Read every tab of each balance workbook. Falls back to STATIC_MEALS if
    a workbook can't be reached at all, and always backfills any tour live
    fetch didn't return (an older tour past its own workbook's window, say)
    from STATIC_MEALS too.
    """
    live: dict = {}
    for key, sheet_id in SHEET_IDS.items():
        data = _fetch_sheet(sheet_id)
        live.update(data)

    if not live:
        print("[meals_sync] Live fetch failed — using static fallback data")
        return dict(STATIC_MEALS)

    # Backfill any tours not returned by live sheets
    for code, meals in STATIC_MEALS.items():
        if code not in live:
            live[code] = meals

    tours_count = len([v for v in live.values() if v])
    print(f"[meals_sync] Loaded meal data for {tours_count} tours")
    return live
