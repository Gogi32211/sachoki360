"""
Track what each tour still owes, read from the BALANCE workbooks. READ-ONLY.

Every tour tab carries its state in three places:

    A1        "ფაზა 1/2/3" — how far the tour's paperwork has got:
              1 = not every invoice has arrived yet
              2 = all invoices in, some still to pay
              3 = all invoices in and paid
    column E  the invoice amount. Filled green once the invoice has arrived and
              been checked; left unfilled while it is still outstanding, so the
              unfilled rows add up to what we have not been billed for yet.
    column F  the payment. Green with at least "ok" in it (often with who paid,
              or whether it was cash) means settled; yellow means still to pay.

Colour is the only marker for most of this, so the fills are read alongside the
values rather than the text alone.
"""
import io
import re
import requests
from openpyxl import load_workbook

from profit_sync import SHEET_IDS, TOUR_CODE_RE, _norm_code, _num, _SUMMARY_KEYS

PHASE_RE = re.compile(r'(?:ფაზა|phase)\s*([123])', re.IGNORECASE)
_OK_RE = re.compile(r'\bok\b', re.IGNORECASE)

AMOUNT_COL = 5   # E
STATUS_COL = 6   # F


def _rgb(cell):
    """(r, g, b) of a cell's fill, or None when it has none."""
    fill = getattr(cell, 'fill', None)
    if fill is None or getattr(fill, 'patternType', None) is None:
        return None
    colour = getattr(fill, 'start_color', None)
    rgb = getattr(colour, 'rgb', None)
    # Theme and indexed colours don't expose an rgb string; treat as unknown.
    if not isinstance(rgb, str) or len(rgb) not in (6, 8):
        return None
    body = rgb[-6:]
    try:
        return tuple(int(body[i:i + 2], 16) for i in (0, 2, 4))
    except ValueError:
        return None


def _shade(cell) -> str:
    """'green', 'yellow', 'none' (unfilled or white) or 'other'.

    Tolerates the range of greens and yellows the sheets actually use, from
    saturated (00B050, FFFF00) through to pale (D9EAD3, FFF2CC): green is what
    leans green, yellow is what drops its blue channel while red and green stay
    high.
    """
    rgb = _rgb(cell)
    if rgb is None:
        return 'none'
    r, g, b = rgb
    if r > 240 and g > 240 and b > 240:
        return 'none'
    if g > r + 12 and g > b + 12:
        return 'green'
    if b < min(r, g) - 20:
        return 'yellow'
    return 'other'


def _phase_from(ws) -> int:
    m = PHASE_RE.search(str(ws['A1'].value or ''))
    return int(m.group(1)) if m else None


def _parse_worksheet(ws):
    """Return (tour_code, debts) or (None, None) when the tab isn't a tour."""
    m = TOUR_CODE_RE.search(ws.title or '')
    code = _norm_code(m.group(1)) if m else None

    out = {
        'phase': _phase_from(ws),
        'invoiced_usd': 0.0,      # every invoice line on the tab
        'received_usd': 0.0,      # invoice in hand and checked (E green)
        'awaited_usd': 0.0,       # not billed to us yet (E unfilled)
        'paid_usd': 0.0,          # settled (F green + "ok")
        'due_usd': 0.0,           # still to pay (F yellow)
        'awaited_count': 0,
        'due_count': 0,
        'lines': [],              # the outstanding ones, for the detail view
    }

    for row in ws.iter_rows(min_col=1, max_col=STATUS_COL):
        cells = list(row)
        if len(cells) < STATUS_COL:
            continue
        name = str(cells[1].value or '').strip()
        if not name or any(name.startswith(k) for k in _SUMMARY_KEYS):
            continue
        if not code:
            mm = TOUR_CODE_RE.search(name)
            if mm:
                code = _norm_code(mm.group(1))
        amount_cell, status_cell = cells[AMOUNT_COL - 1], cells[STATUS_COL - 1]
        usd = _num(amount_cell.value)
        if not usd or usd <= 0:
            continue

        e_shade = _shade(amount_cell)
        f_shade = _shade(status_cell)
        status_text = str(status_cell.value or '').strip()

        # "ok" in a green F cell means settled — that is the reliable marker.
        paid = f_shade == 'green' and bool(_OK_RE.search(status_text))
        # An invoice that has been paid was necessarily received, whatever colour
        # E carries: older tours predate the convention of greening it.
        received = paid or e_shade == 'green'
        due = (not paid) and (f_shade == 'yellow' or received)

        out['invoiced_usd'] += usd
        if received:
            out['received_usd'] += usd
        else:
            out['awaited_usd'] += usd
            out['awaited_count'] += 1
        if paid:
            out['paid_usd'] += usd
        elif due:
            out['due_usd'] += usd
            out['due_count'] += 1

        if not paid:
            out['lines'].append({
                'name': name,
                'usd': round(usd, 2),
                'received': received,
                'note': status_text,
            })

    if not code:
        return None, None

    # A1 states where the tour stands, and it outranks the cell colouring: a tour
    # marked phase 2 or 3 has every invoice in hand even where E was never
    # greened, and phase 3 is settled in full.
    if out['phase'] in (2, 3):
        out['received_usd'] = out['invoiced_usd']
        out['awaited_usd'] = 0.0
        out['awaited_count'] = 0
        for line in out['lines']:
            line['received'] = True
    if out['phase'] == 3:
        out['paid_usd'] = out['invoiced_usd']
        out['due_usd'] = 0.0
        out['due_count'] = 0
        out['lines'] = []

    for k in ('invoiced_usd', 'received_usd', 'awaited_usd', 'paid_usd', 'due_usd'):
        out[k] = round(out[k], 2)
    out['lines'].sort(key=lambda l: -l['usd'])
    return code, out


def _fetch_workbook(sheet_id: str) -> dict:
    url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"
    results: dict = {}
    try:
        resp = requests.get(url, timeout=90)
        resp.raise_for_status()
        # Styles are the point here, so this can't run with values_only.
        wb = load_workbook(io.BytesIO(resp.content), data_only=True, read_only=True)
        for ws in wb.worksheets:
            code, data = _parse_worksheet(ws)
            if code and data:
                results[code] = data
        wb.close()
        print(f"[debts_sync] {sheet_id}: parsed {len(results)} tours")
    except Exception as e:
        print(f"[debts_sync] Could not fetch/parse {sheet_id}: {e}")
    return results


def fetch_tour_debts() -> dict:
    """Return {tour_code: {phase, invoiced_usd, awaited_usd, due_usd, ...}}."""
    results: dict = {}
    for sheet_id in SHEET_IDS.values():
        results.update(_fetch_workbook(sheet_id))
    print(f"[debts_sync] Total: {len(results)} tours")
    return results
