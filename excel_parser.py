from datetime import date, datetime
import openpyxl
import os
import re

EXCEL_FILES = [
    "/root/.claude/uploads/24a9ba42-2faf-5339-ae9b-5a5487a67c79/af288247-LN_TOURS_BALANCE.xlsx",
    "/root/.claude/uploads/24a9ba42-2faf-5339-ae9b-5a5487a67c79/78a04bcd-ZT_TOURS_BALANCE.xlsx",
    "/root/.claude/uploads/24a9ba42-2faf-5339-ae9b-5a5487a67c79/6071171f-KT_DT2_TOURS_BALANCE.xlsx",
]

def parse_all_excel():
    """Returns list of dicts: {tour_code, guide, rooms, meals: [...], financials: {...}}"""
    results = []
    for fpath in EXCEL_FILES:
        if not os.path.exists(fpath):
            continue
        wb = openpyxl.load_workbook(fpath, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            parsed = parse_sheet(ws)
            if parsed:
                results.append(parsed)
    return results

def parse_sheet(ws):
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return None

    # Get tour code from row 2 col B (index 1)
    tour_code = None
    guide = None
    rooms = None
    for row in rows[:5]:
        if len(row) > 1 and row[1] and isinstance(row[1], str) and '-' in row[1]:
            val = row[1].strip()
            # Match pattern like "LN-0608", "ZT-0615", "KT-0609", "DT1-0531", "DT2-0511"
            if re.match(r'^(ZT|LN|KT|DT1|DT2)-\d{4}$', val):
                tour_code = val
                guide = row[2] if len(row) > 2 else None
                rooms = row[3] if len(row) > 3 else None
                break

    if not tour_code:
        return None

    meals = []
    financials = {
        'spent_gel': 0, 'tour_price_gel': 0, 'tour_price_usd': 0,
        'profit_gel': 0, 'paid_gel': 0, 'paid_cny': 0, 'due_gel': 0,
    }

    in_meals = False

    for row in rows:
        if len(row) < 5:
            continue

        col_a = row[0]
        col_b = row[1] if len(row) > 1 else None
        col_d = row[3] if len(row) > 3 else None
        col_e = row[4] if len(row) > 4 else None
        col_f = row[5] if len(row) > 5 else None

        # Detect start of meals section: a cell like "LN0608/ 17+1" or "ZT-0615 / 20+1"
        if col_b and isinstance(col_b, str) and '/' in col_b:
            stripped = col_b.replace(' ', '').replace('-', '')
            tc_stripped = tour_code.replace(' ', '').replace('-', '')
            if tc_stripped in stripped:
                in_meals = True
                continue

        # Parse meal rows
        if in_meals and isinstance(col_a, datetime):
            desc = str(col_b or '').strip()
            meal_type = None
            restaurant = None

            if 'ლანჩი' in desc:
                meal_type = 'lunch'
                restaurant = desc.split('-', 1)[1].strip() if '-' in desc else desc.replace('ლანჩი', '').strip()
            elif 'ვახშამი' in desc or 'ვაშამი' in desc:
                meal_type = 'dinner'
                restaurant = desc.split('-', 1)[1].strip() if '-' in desc else desc.replace('ვახშამი', '').strip()
            elif 'დეგუსტაცია' in desc:
                meal_type = 'degustation'
                restaurant = desc.replace('დეგუსტაცია', '').strip()

            if meal_type and restaurant:
                gel = col_d if isinstance(col_d, (int, float)) else 0
                usd = col_e if isinstance(col_e, (int, float)) else 0
                meals.append({
                    'date': col_a.date().isoformat(),
                    'meal_type': meal_type,
                    'restaurant': restaurant,
                    'gel_amount': round(float(gel), 2),
                    'usd_amount': round(float(usd), 2),
                })

        # Financial summary rows: label is in col D (index 3)
        if isinstance(col_d, str):
            label = col_d.strip()
            gel_val = col_e if isinstance(col_e, (int, float)) else 0
            usd_val = col_f if isinstance(col_f, (int, float)) else 0

            if label == 'დაიხარჯა':
                financials['spent_gel'] = round(float(gel_val), 2)
            elif label == 'ტურის ღირებ.':
                financials['tour_price_gel'] = round(float(gel_val), 2)
                financials['tour_price_usd'] = round(float(usd_val), 2)
            elif label == 'მოგება':
                financials['profit_gel'] = round(float(gel_val), 2)
            elif label == 'ჩარიცხული':
                financials['paid_gel'] = round(float(gel_val), 2)
                financials['paid_cny'] = round(float(usd_val), 2)
            elif label == 'ჩასარიცხია':
                financials['due_gel'] = round(float(gel_val), 2)

    return {
        'tour_code': tour_code,
        'guide': str(guide).strip() if guide else '',
        'rooms': str(rooms).strip() if rooms else '',
        'meals': meals,
        'financials': financials,
    }
